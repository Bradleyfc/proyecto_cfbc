from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView
from django.db.models import Q, Prefetch
from django.contrib.auth.models import User
from .models import (
    Curso, SolicitudInscripcion, RespuestaEstudiante, 
    FormularioAplicacion, PreguntaFormulario, OpcionRespuesta,
    CursoAcademico
)

def es_profesor_o_secretaria(user):
    """Verifica si el usuario es profesor o secretaria"""
    return user.groups.filter(name__in=['Profesores', 'Secretaria']).exists()

class RegistroRespuestasGeneralView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Vista para mostrar un registro general de todas las respuestas de formularios
    organizadas por curso y estudiante
    """
    model = Curso
    template_name = 'registro_respuestas/registro_general.html'
    context_object_name = 'cursos'
    
    def test_func(self):
        return es_profesor_o_secretaria(self.request.user)
    
    def get_queryset(self):
        # Obtener el curso académico activo
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        
        # Filtrar cursos que tengan formularios de aplicación y solicitudes
        queryset = Curso.objects.filter(
            formulario_aplicacion__isnull=False,
            solicitudes__isnull=False,
            curso_academico=curso_academico_activo
        ).distinct().prefetch_related(
            'formulario_aplicacion',
            Prefetch(
                'solicitudes',
                queryset=SolicitudInscripcion.objects.select_related('estudiante').prefetch_related(
                    Prefetch(
                        'respuestas',
                        queryset=RespuestaEstudiante.objects.select_related('pregunta').prefetch_related('opciones_seleccionadas')
                    )
                )
            )
        )
        
        # Filtros opcionales
        search = self.request.GET.get('search')
        estado_filter = self.request.GET.get('estado')
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(teacher__first_name__icontains=search) |
                Q(teacher__last_name__icontains=search)
            )
        
        if estado_filter:
            queryset = queryset.filter(solicitudes__estado=estado_filter)
        
        return queryset.order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Agregar estadísticas generales
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        context['curso_academico'] = curso_academico_activo
        
        if curso_academico_activo:
            context['total_solicitudes'] = SolicitudInscripcion.objects.filter(
                curso__curso_academico=curso_academico_activo
            ).count()
            
            context['solicitudes_pendientes'] = SolicitudInscripcion.objects.filter(
                curso__curso_academico=curso_academico_activo,
                estado='pendiente'
            ).count()
            
            context['solicitudes_aprobadas'] = SolicitudInscripcion.objects.filter(
                curso__curso_academico=curso_academico_activo,
                estado='aprobada'
            ).count()
            
            context['solicitudes_rechazadas'] = SolicitudInscripcion.objects.filter(
                curso__curso_academico=curso_academico_activo,
                estado='rechazada'
            ).count()
        
        # Agregar estadísticas por curso
        cursos = context['cursos']
        for curso in cursos:
            curso.total_solicitudes = curso.solicitudes.count()
            curso.solicitudes_pendientes = curso.solicitudes.filter(estado='pendiente').count()
            curso.solicitudes_aprobadas = curso.solicitudes.filter(estado='aprobada').count()
            curso.solicitudes_rechazadas = curso.solicitudes.filter(estado='rechazada').count()
        
        # Opciones para filtros
        context['estados_choices'] = SolicitudInscripcion.ESTADO_CHOICES
        
        return context

class RegistroRespuestasCursoView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """
    Vista para mostrar las respuestas de un curso específico
    organizadas por estudiante
    """
    model = Curso
    template_name = 'registro_respuestas/registro_curso.html'
    context_object_name = 'curso'
    
    def test_func(self):
        return es_profesor_o_secretaria(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso = self.get_object()
        
        # Obtener todas las solicitudes del curso con sus respuestas
        solicitudes = SolicitudInscripcion.objects.filter(
            curso=curso
        ).select_related('estudiante').prefetch_related(
            Prefetch(
                'respuestas',
                queryset=RespuestaEstudiante.objects.select_related('pregunta').prefetch_related('opciones_seleccionadas')
            )
        ).order_by('estudiante__first_name', 'estudiante__last_name')
        
        # Filtros opcionales
        estado_filter = self.request.GET.get('estado')
        search = self.request.GET.get('search')
        
        if estado_filter:
            solicitudes = solicitudes.filter(estado=estado_filter)
        
        if search:
            solicitudes = solicitudes.filter(
                Q(estudiante__first_name__icontains=search) |
                Q(estudiante__last_name__icontains=search) |
                Q(estudiante__username__icontains=search)
            )
        
        context['solicitudes'] = solicitudes
        
        # Obtener el formulario del curso para mostrar las preguntas
        try:
            formulario = FormularioAplicacion.objects.get(curso=curso)
            preguntas = formulario.preguntas.all().order_by('orden')
            context['formulario'] = formulario
            context['preguntas'] = preguntas
        except FormularioAplicacion.DoesNotExist:
            context['formulario'] = None
            context['preguntas'] = []
        
        # Estadísticas del curso
        context['total_solicitudes_curso'] = solicitudes.count()
        context['solicitudes_pendientes_curso'] = solicitudes.filter(estado='pendiente').count()
        context['solicitudes_aprobadas_curso'] = solicitudes.filter(estado='aprobada').count()
        context['solicitudes_rechazadas_curso'] = solicitudes.filter(estado='rechazada').count()
        
        # Opciones para filtros
        context['estados_choices'] = SolicitudInscripcion.ESTADO_CHOICES
        
        return context

class RegistroRespuestasEstudianteView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """
    Vista para mostrar las respuestas detalladas de un estudiante específico
    """
    model = SolicitudInscripcion
    template_name = 'registro_respuestas/registro_estudiante.html'
    context_object_name = 'solicitud'
    
    def test_func(self):
        return es_profesor_o_secretaria(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        solicitud = self.get_object()
        
        # Obtener todas las respuestas del estudiante para esta solicitud
        respuestas = RespuestaEstudiante.objects.filter(
            solicitud=solicitud
        ).select_related('pregunta').prefetch_related('opciones_seleccionadas').order_by('pregunta__orden')
        
        context['respuestas'] = respuestas
        
        # Obtener información del registro del estudiante
        try:
            from accounts.models import Registro
            registro = Registro.objects.get(user=solicitud.estudiante)
            context['registro_estudiante'] = registro
        except Registro.DoesNotExist:
            context['registro_estudiante'] = None
        
        return context

@login_required
@user_passes_test(es_profesor_o_secretaria)
def exportar_respuestas_excel(request, curso_id=None):
    """
    Vista para exportar las respuestas a Excel
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from io import BytesIO
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Estilos
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    if curso_id:
        # Exportar respuestas de un curso específico
        curso = get_object_or_404(Curso, id=curso_id)
        ws.title = f"Respuestas {curso.name}"
        
        solicitudes = SolicitudInscripcion.objects.filter(
            curso=curso
        ).select_related('estudiante').prefetch_related(
            'respuestas__pregunta',
            'respuestas__opciones_seleccionadas'
        ).order_by('estudiante__first_name', 'estudiante__last_name')
        
        # Obtener preguntas del formulario
        try:
            formulario = FormularioAplicacion.objects.get(curso=curso)
            preguntas = list(formulario.preguntas.all().order_by('orden'))
        except FormularioAplicacion.DoesNotExist:
            preguntas = []
        
        # Encabezados
        headers = ['Estudiante', 'Email', 'Estado', 'Fecha Solicitud']
        for pregunta in preguntas:
            headers.append(pregunta.texto)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, solicitud in enumerate(solicitudes, 2):
            ws.cell(row=row_num, column=1, value=solicitud.estudiante.get_full_name() or solicitud.estudiante.username)
            ws.cell(row=row_num, column=2, value=solicitud.estudiante.email)
            ws.cell(row=row_num, column=3, value=solicitud.get_estado_display())
            ws.cell(row=row_num, column=4, value=solicitud.fecha_solicitud.strftime('%d/%m/%Y %H:%M'))
            
            # Respuestas por pregunta
            respuestas_dict = {}
            for respuesta in solicitud.respuestas.all():
                opciones_texto = ', '.join([opcion.texto for opcion in respuesta.opciones_seleccionadas.all()])
                respuestas_dict[respuesta.pregunta.id] = opciones_texto
            
            for col_num, pregunta in enumerate(preguntas, 5):
                respuesta_texto = respuestas_dict.get(pregunta.id, 'Sin respuesta')
                ws.cell(row=row_num, column=col_num, value=respuesta_texto)
            
            # Aplicar bordes
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = border
        
        filename = f"respuestas_{curso.name.replace(' ', '_')}.xlsx"
    
    else:
        # Exportar resumen general de todos los cursos
        ws.title = "Resumen General"
        
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        cursos = Curso.objects.filter(
            formulario_aplicacion__isnull=False,
            solicitudes__isnull=False,
            curso_academico=curso_academico_activo
        ).distinct()
        
        # Encabezados
        headers = ['Curso', 'Profesor', 'Total Solicitudes', 'Pendientes', 'Aprobadas', 'Rechazadas']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, curso in enumerate(cursos, 2):
            total = curso.solicitudes.count()
            pendientes = curso.solicitudes.filter(estado='pendiente').count()
            aprobadas = curso.solicitudes.filter(estado='aprobada').count()
            rechazadas = curso.solicitudes.filter(estado='rechazada').count()
            
            ws.cell(row=row_num, column=1, value=curso.name)
            ws.cell(row=row_num, column=2, value=curso.teacher.get_full_name() or curso.teacher.username)
            ws.cell(row=row_num, column=3, value=total)
            ws.cell(row=row_num, column=4, value=pendientes)
            ws.cell(row=row_num, column=5, value=aprobadas)
            ws.cell(row=row_num, column=6, value=rechazadas)
            
            # Aplicar bordes
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = border
        
        filename = "resumen_respuestas_general.xlsx"
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
        ws.column_dimensions[column].width = adjusted_width
    
    # Guardar archivo
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response