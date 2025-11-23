from typing import override
from django.contrib.auth.forms import UserCreationForm
from django.views import View
from django.core.mail import send_mail
from django.conf import settings
from django.views.generic import ListView, DetailView, TemplateView, CreateView, UpdateView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.urls import reverse_lazy, reverse
from django.utils.decorators import method_decorator
from django.contrib.auth import logout
from django.contrib import messages
from django.utils import timezone
from .forms import (
    CustomUserCreationForm, CourseForm, CalificacionesForm, NotaIndividualFormSet,
    FormularioAplicacionForm, PreguntaFormularioForm, OpcionRespuestaForm,
    OpcionRespuestaFormSet, PreguntaFormularioFormSet, RespuestaEstudianteForm
)
from django.contrib.auth.models import Group, User
from django.db.models import Q, Max
from datetime import date, datetime
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from accounts.models import Registro
from blog.models import Noticia
from .models import (
    CursoAcademico, Curso, Matriculas, Calificaciones, Asistencia,
    FormularioAplicacion, PreguntaFormulario, OpcionRespuesta, SolicitudInscripcion, RespuestaEstudiante
)

# Create your views here.

class UsuariosRegistradosView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Registro
    template_name = 'usuarios_registrados.html'
    context_object_name = 'registros'

    def test_func(self):
        return self.request.user.groups.filter(name='Secretaria').exists()

    def get_queryset(self):
        queryset = Registro.objects.all().select_related('user').prefetch_related('user__groups')
        # Eliminar o comentar las siguientes líneas si ya no necesitas el filtro por grupo desde la URL
        # grupo = self.request.GET.get('grupo')
        search = self.request.GET.get('search')

        # Asegurarse de que solo se muestren los estudiantes
        queryset = queryset.filter(user__groups__name='Estudiantes')

        # if grupo:
        #     queryset = queryset.filter(user__groups__name=grupo)

        if search:
            queryset = queryset.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search)
            )

        return queryset


@login_required
def export_usuarios_excel(request):
    search_query = request.GET.get('search', '')
    registros = Registro.objects.filter(user__groups__name='Estudiantes')

    if search_query:
        registros = registros.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(carnet__icontains=search_query)
        ).distinct()

    context = {
        'registros': registros
    }
    excel_file = generate_excel(context)
    response = HttpResponse(excel_file.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="usuarios_registrados.xlsx"'
    return response

@login_required
def export_matriculas_pdf(request):
    curso_academico_id = request.GET.get('curso_academico')
    curso_id = request.GET.get('curso')
    student_id = request.GET.get('student')

    matriculas = Matriculas.objects.all()

    if curso_academico_id:
        matriculas = matriculas.filter(course__curso_academico__id=curso_academico_id)
    if curso_id:
        matriculas = matriculas.filter(course__id=curso_id)
    if student_id:
        matriculas = matriculas.filter(student__id=student_id)

    context = {
        'matriculas': matriculas,
        'curso_academico': CursoAcademico.objects.get(id=curso_academico_id) if curso_academico_id else None,
    }
    return render_to_pdf('matriculas_pdf.html', context)

@login_required
def export_matriculas_excel(request):
    curso_academico_id = request.GET.get('curso_academico')
    curso_id = request.GET.get('curso')
    student_id = request.GET.get('student')

    matriculas = Matriculas.objects.all()

    if curso_academico_id:
        matriculas = matriculas.filter(course__curso_academico__id=curso_academico_id)
    if curso_id:
        matriculas = matriculas.filter(course__id=curso_id)
    if student_id:
        matriculas = matriculas.filter(student__id=student_id)

    context = {
        'matriculas': matriculas,
        'curso_academico': CursoAcademico.objects.get(id=curso_academico_id) if curso_academico_id else None,
    }
    
    # Generar el archivo Excel
    output = generate_excel(context)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=matriculas.xlsx'
    return response

# Función auxiliar para generar PDF
def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

# Función auxiliar para generar Excel
def generate_excel(context_dict={}):
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    
    # Estilos para el Excel
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Obtener datos del contexto
    curso_academico = context_dict.get('curso_academico')
    cursos = context_dict.get('cursos', [])
    matriculas = context_dict.get('matriculas', [])
    calificaciones = context_dict.get('calificaciones', [])
    asistencias = context_dict.get('asistencias', [])
    registros = context_dict.get('registros', []) # Añadir registros al contexto
    
    # Hoja de información general
    if curso_academico:
        ws_info = wb.active
        ws_info.title = "Información General"
        ws_info['A1'] = f"Curso Académico: {curso_academico.nombre}"
        ws_info['A2'] = f"Activo: {'Sí' if curso_academico.activo else 'No'}"
        ws_info['A3'] = f"Archivado: {'Sí' if curso_academico.archivado else 'No'}"
        ws_info['A4'] = f"Fecha de Creación: {curso_academico.fecha_creacion}"
    else:
        # Remove the default active sheet if no curso_academico is provided
        if 'Sheet' in wb.sheetnames:
            std = wb['Sheet']
            wb.remove(std)
    
    # Hoja de cursos
    if cursos:
        ws_cursos = wb.create_sheet(title="Cursos")
        # Encabezados
        headers = ["Nombre del Curso", "Profesor", "Estado"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_cursos.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, curso in enumerate(cursos, 2):
            ws_cursos.cell(row=row_num, column=1, value=curso.name)
            ws_cursos.cell(row=row_num, column=2, value=curso.teacher.get_full_name() or curso.teacher.username)
            ws_cursos.cell(row=row_num, column=3, value=curso.get_status_display())
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, 4):
                ws_cursos.cell(row=row_num, column=col_num).border = border
        
        # Ajustar ancho de columnas
        for col in ws_cursos.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_cursos.column_dimensions[column].width = adjusted_width
    
    # Hoja de matrículas
    if matriculas:
        ws_matriculas = wb.create_sheet(title="Matrículas")
        # Encabezados
        headers = ["Estudiante", "Curso Académico", "Curso", "Fecha Matrícula", "Estado Matrícula"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_matriculas.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, matricula in enumerate(matriculas, 2):
            ws_matriculas.cell(row=row_num, column=1, value=matricula.student.get_full_name() or matricula.student.username)
            ws_matriculas.cell(row=row_num, column=2, value=matricula.course.curso_academico.nombre if matricula.course and matricula.course.curso_academico else 'N/A')
            ws_matriculas.cell(row=row_num, column=3, value=matricula.course.name if matricula.course else 'N/A')
            ws_matriculas.cell(row=row_num, column=4, value=matricula.fecha_matricula.strftime('%d/%m/%Y') if matricula.fecha_matricula else 'N/A')
            ws_matriculas.cell(row=row_num, column=5, value=matricula.get_estado_display())
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, 6):
                ws_matriculas.cell(row=row_num, column=col_num).border = border
        
        # Ajustar ancho de columnas
        for col in ws_matriculas.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_matriculas.column_dimensions[column].width = adjusted_width

    
    # Hoja de calificaciones
    if calificaciones:
        ws_calificaciones = wb.create_sheet(title="Calificaciones")
        
        # Determinar el número máximo de notas individuales
        max_notas = 0
        for calificacion in calificaciones:
            num_notas = calificacion.notas.count()
            if num_notas > max_notas:
                max_notas = num_notas
        
        # Crear encabezados dinámicos
        headers = ["Estudiante", "Curso"]
        for i in range(1, max_notas + 1):
            headers.append(f"Nota {i}")
        headers.append("Promedio")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_calificaciones.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, calificacion in enumerate(calificaciones, 2):
            # Información básica
            ws_calificaciones.cell(row=row_num, column=1, value=calificacion.student.get_full_name() or calificacion.student.username)
            ws_calificaciones.cell(row=row_num, column=2, value=calificacion.course.name)
            
            # Notas individuales
            notas_individuales = list(calificacion.notas.all().order_by('fecha_creacion'))
            for i, nota in enumerate(notas_individuales, 3):
                ws_calificaciones.cell(row=row_num, column=i, value=nota.valor)
            
            # Rellenar con N/A las notas que no existen
            for i in range(len(notas_individuales) + 3, max_notas + 3):
                ws_calificaciones.cell(row=row_num, column=i, value='N/A')
            
            # Promedio
            ws_calificaciones.cell(row=row_num, column=max_notas + 3, value=calificacion.average if calificacion.average is not None else 'N/A')
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, max_notas + 4):
                ws_calificaciones.cell(row=row_num, column=col_num).border = border
        
    # Hoja de usuarios registrados
    if registros:
        ws_usuarios = wb.create_sheet(title="Usuarios Registrados")
        # Encabezados
        headers = [
            "Nombre", "Apellidos", "Email", "Nacionalidad", "Carnet ID", "Carnet Disponible", "Sexo",
            "Dirección", "Municipio", "Provincia", "Movil", "Grado Académico",
            "Ocupación", "Título", "Título Disponible", "Grupo", "Fecha de Registro"
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws_usuarios.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, registro in enumerate(registros, 2):
            ws_usuarios.cell(row=row_num, column=1, value=registro.user.first_name)
            ws_usuarios.cell(row=row_num, column=2, value=registro.user.last_name)
            ws_usuarios.cell(row=row_num, column=3, value=registro.user.email)
            ws_usuarios.cell(row=row_num, column=4, value=registro.nacionalidad)
            ws_usuarios.cell(row=row_num, column=5, value=registro.carnet)
            ws_usuarios.cell(row=row_num, column=6, value="Sí" if registro.foto_carnet else "No")
            ws_usuarios.cell(row=row_num, column=7, value=registro.sexo)
            ws_usuarios.cell(row=row_num, column=8, value=registro.address)
            ws_usuarios.cell(row=row_num, column=9, value=registro.location)
            ws_usuarios.cell(row=row_num, column=10, value=registro.provincia)
            ws_usuarios.cell(row=row_num, column=11, value=registro.movil)
            ws_usuarios.cell(row=row_num, column=12, value=registro.get_grado_display())
            ws_usuarios.cell(row=row_num, column=13, value=registro.get_ocupacion_display())
            ws_usuarios.cell(row=row_num, column=14, value=registro.titulo)
            ws_usuarios.cell(row=row_num, column=15, value="Sí" if registro.foto_titulo else "No")
            ws_usuarios.cell(row=row_num, column=16, value=registro.user.groups.first().name if registro.user.groups.first() else '')
            ws_usuarios.cell(row=row_num, column=17, value=registro.user.date_joined.strftime("%d/%m/%Y"))
            
            for col_num in range(1, len(headers) + 1):
                ws_usuarios.cell(row=row_num, column=col_num).border = border
        
        # Ajustar ancho de columnas
        for col in ws_usuarios.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_usuarios.column_dimensions[column].width = adjusted_width

    if calificaciones:
        # Ajustar ancho de columnas
        for col in ws_calificaciones.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_calificaciones.column_dimensions[column].width = adjusted_width

    # Hoja de asistencias
    if asistencias:
        ws_asistencias = wb.create_sheet(title="Asistencias")
        # Encabezados
        headers = ["Estudiante", "Curso", "Fecha", "Presente"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_asistencias.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Datos
        for row_num, asistencia in enumerate(asistencias, 2):
            ws_asistencias.cell(row=row_num, column=1, value=asistencia.student.get_full_name() or asistencia.student.username)
            ws_asistencias.cell(row=row_num, column=2, value=asistencia.course.name)
            ws_asistencias.cell(row=row_num, column=3, value=asistencia.date.strftime('%d/%m/%Y') if asistencia.date else 'N/A')
            ws_asistencias.cell(row=row_num, column=4, value='Sí' if asistencia.presente else 'No')

            # Aplicar bordes a todas las celdas
            for col_num in range(1, 5):
                ws_asistencias.cell(row=row_num, column=col_num).border = border

        # Ajustar ancho de columnas
        for col in ws_asistencias.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws_asistencias.column_dimensions[column].width = adjusted_width

    # Guardar el archivo en memoria
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file

class CursoAcademicoDetailView(DetailView):
    model = CursoAcademico
    template_name = 'curso_academico_detail.html'
    context_object_name = 'curso_academico'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso_academico = self.get_object()
        
        # Obtener los filtros de la URL
        curso_id = self.request.GET.get('curso')
        estudiante_id = self.request.GET.get('estudiante')
        
        # Filtrar cursos
        cursos = Curso.objects.filter(matriculas__curso_academico=curso_academico).distinct()
        if curso_id:
            cursos = cursos.filter(id=curso_id)
        context['cursos'] = cursos
        
        # Filtrar matrículas
        matriculas = Matriculas.objects.filter(curso_academico=curso_academico)
        if curso_id:
            matriculas = matriculas.filter(course_id=curso_id)
        if estudiante_id:
            matriculas = matriculas.filter(student_id=estudiante_id)
        context['matriculas'] = matriculas
        
        # Filtrar calificaciones
        calificaciones = Calificaciones.objects.filter(curso_academico=curso_academico)
        if curso_id:
            calificaciones = calificaciones.filter(course_id=curso_id)
        if estudiante_id:
            calificaciones = calificaciones.filter(student_id=estudiante_id)
        context['calificaciones'] = calificaciones
        
        # Filtrar asistencias
        asistencias = Asistencia.objects.filter(course__matriculas__curso_academico=curso_academico).distinct()
        if curso_id:
            asistencias = asistencias.filter(course_id=curso_id)
        if estudiante_id:
            asistencias = asistencias.filter(student_id=estudiante_id)
        context['asistencias'] = asistencias
        
        # Agregar listas para los selectores de filtro
        context['cursos_disponibles'] = Curso.objects.filter(matriculas__curso_academico=curso_academico).distinct()
        context['estudiantes_disponibles'] = User.objects.filter(
            matriculas__curso_academico=curso_academico
        ).distinct()
        
        return context
        
    def render_to_response(self, context, **response_kwargs):
        # Verificar si se solicita PDF
        if 'pdf' in self.request.GET:
            # Añadir fecha actual al contexto para el PDF
            context['now'] = datetime.now()
            # Crear un template para PDF basado en el mismo contexto
            pdf = render_to_pdf('curso_academico_pdf.html', context)
            if pdf:
                response = HttpResponse(pdf, content_type='application/pdf')
                filename = f"curso_academico_{context['curso_academico'].nombre}.pdf"
                content = f"attachment; filename={filename}"
                response['Content-Disposition'] = content
                return response
        # Verificar si se solicita Excel
        elif 'excel' in self.request.GET:
            # Generar archivo Excel
            excel_file = generate_excel(context)
            if excel_file:
                response = HttpResponse(excel_file.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                filename = f"curso_academico_{context['curso_academico'].nombre}.xlsx"
                response['Content-Disposition'] = f'attachment; filename={filename}'
                return response
        # Si no se solicita PDF ni Excel, renderizar normalmente
        return super().render_to_response(context, **response_kwargs)


class BaseContextMixin:
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        group_name = None
        if user.is_authenticated:
            group = Group.objects.filter(user=user).first()
            if group:
                group_name = group.name
            context['group_name'] = group_name
        return context


class HomeView(BaseContextMixin, TemplateView):
    template_name = 'home.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        if curso_academico_activo:
            courses = Curso.objects.filter(curso_academico=curso_academico_activo)
        else:
            courses = Curso.objects.none()
        
        # Obtener todos los formularios de aplicación existentes
        formularios = FormularioAplicacion.objects.all()
        formularios_por_curso = {f.curso_id: f for f in formularios}
        
        # Asignar los formularios a los cursos
        for curso in courses:
            if curso.id in formularios_por_curso:
                curso.formulario_aplicacion = formularios_por_curso[curso.id]
            else:
                curso.formulario_aplicacion = None
        
        # Group courses into chunks of four for the carousel
        grouped_courses = [courses[i:i + 4] for i in range(0, len(courses), 4)]
        context['grouped_courses'] = grouped_courses
        
        # Obtener las noticias publicadas más recientes
        noticias = Noticia.objects.filter(estado='publicado').order_by('-fecha_publicacion')[:8]
        
        # Agrupar noticias en chunks de 4 para el carousel
        grouped_noticias = [noticias[i:i + 4] for i in range(0, len(noticias), 4)]
        context['grouped_noticias'] = grouped_noticias
        student = self.request.user if self.request.user.is_authenticated else None
        
        # Crear conjuntos de IDs de cursos con solicitudes pendientes y rechazadas
        cursos_con_solicitudes_pendientes = set()
        cursos_con_solicitudes_rechazadas = set()
        
        if student:
            # Obtener todos los IDs de cursos con solicitudes pendientes
            cursos_con_solicitudes_pendientes = set(
                SolicitudInscripcion.objects.filter(
                    estudiante=student,
                    estado='pendiente'
                ).values_list('curso_id', flat=True)
            )
            
            # Obtener todos los IDs de cursos con solicitudes rechazadas
            cursos_con_solicitudes_rechazadas = set(
                SolicitudInscripcion.objects.filter(
                    estudiante=student,
                    estado='rechazada'
                ).values_list('curso_id', flat=True)
            )
            
            print(f"DEBUG: Home - Cursos con solicitudes pendientes: {cursos_con_solicitudes_pendientes}")
            print(f"DEBUG: Home - Cursos con solicitudes rechazadas: {cursos_con_solicitudes_rechazadas}")

        for item in courses:
            if student:
                # Verificar si el estudiante está matriculado
                item.is_enrolled = Matriculas.objects.filter(
                    course=item, 
                    student=student
                ).exists()
                
                # Verificar si el estudiante tiene una solicitud pendiente para este curso
                item.tiene_solicitud_pendiente = item.id in cursos_con_solicitudes_pendientes
                
                # Verificar si el estudiante tiene una solicitud rechazada para este curso
                item.tiene_solicitud_rechazada = item.id in cursos_con_solicitudes_rechazadas
                
                if item.tiene_solicitud_pendiente:
                    print(f"DEBUG: Home - Curso {item.name} (ID: {item.id}) tiene solicitud pendiente")
                if item.tiene_solicitud_rechazada:
                    print(f"DEBUG: Home - Curso {item.name} (ID: {item.id}) tiene solicitud rechazada")
            else:
                item.is_enrolled = False
                item.tiene_solicitud_pendiente = False
                item.tiene_solicitud_rechazada = False
            
            # Calcular el conteo de inscripciones
            enrollment_count = Matriculas.objects.filter(course=item).count()
            item.enrollment_count = enrollment_count
            
            # Ya no necesitamos este código porque los formularios se cargan en get_context_data

        context['courses'] = courses
        return context
        return context



class ListadoCursosView(BaseContextMixin, ListView):
    model = Curso
    template_name = 'cursos.html'
    context_object_name = 'courses'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        # Obtener el CursoAcademico activo
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        if curso_academico_activo:
            # Filtrar los cursos por el CursoAcademico activo
            return Curso.objects.filter(curso_academico=curso_academico_activo)
        return Curso.objects.none() # No mostrar cursos si no hay un CursoAcademico activo
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.request.user if self.request.user.is_authenticated else None
        
        # Crear conjuntos de IDs de cursos con solicitudes pendientes y rechazadas
        cursos_con_solicitudes_pendientes = set()
        cursos_con_solicitudes_rechazadas = set()
        
        if student:
            # Obtener todos los IDs de cursos con solicitudes pendientes
            cursos_con_solicitudes_pendientes = set(
                SolicitudInscripcion.objects.filter(
                    estudiante=student,
                    estado='pendiente'
                ).values_list('curso_id', flat=True)
            )
            
            # Obtener todos los IDs de cursos con solicitudes rechazadas
            cursos_con_solicitudes_rechazadas = set(
                SolicitudInscripcion.objects.filter(
                    estudiante=student,
                    estado='rechazada'
                ).values_list('curso_id', flat=True)
            )
        
        # Obtener todos los formularios de aplicación existentes
        formularios = FormularioAplicacion.objects.all()
        formularios_por_curso = {f.curso_id: f for f in formularios}
        
        # Procesar cada curso
        for course in context['courses']:
            if student:
                # Verificar si el estudiante está matriculado
                course.is_enrolled = Matriculas.objects.filter(
                    course=course, 
                    student=student
                ).exists()
                
                # Verificar si el estudiante tiene una solicitud pendiente para este curso
                course.tiene_solicitud_pendiente = course.id in cursos_con_solicitudes_pendientes
                
                # Verificar si el estudiante tiene una solicitud rechazada para este curso
                course.tiene_solicitud_rechazada = course.id in cursos_con_solicitudes_rechazadas
            else:
                course.is_enrolled = False
                course.tiene_solicitud_pendiente = False
                course.tiene_solicitud_rechazada = False
            
            # Calcular el conteo de inscripciones
            enrollment_count = Matriculas.objects.filter(course=course).count()
            course.enrollment_count = enrollment_count
            
            # Asignar el formulario de aplicación si existe
            if course.id in formularios_por_curso:
                course.formulario_aplicacion = formularios_por_curso[course.id]
            else:
                course.formulario_aplicacion = None
        
        return context
            
        return context

# para cerrar sesion


def logout_view(request):
    logout(request)
    return redirect('principal:home')


# pagina de Registro
""" class RegisterView(View):
    def get(self, request):
        data = {
            'form': RegisterForm()
        }
        return render(request, 'registration/registro.html', data)

    def post(self, request):
        user_creation_form=RegisterForm(data=request.POST)
        if user_creation_form.is_valid():
            user_creation_form.save()  
            user = authenticate(username=user_creation_form.cleaned_data['username'],
                                password=user_creation_form.cleaned_data['password'])
            login(request, user)
            return redirect('principal:home')
        data = {
            'form':user_creation_form
        }   
        return render(request, 'registration/registro.html', data) """


import random

def registro(request):
    if request.method == 'POST':
        user_creation_form = CustomUserCreationForm(
            data=request.POST, files=request.FILES)

        if user_creation_form.is_valid():
            # Generar código aleatorio de 4 dígitos
            verification_code = str(random.randint(1000, 9999))
            # Almacenar datos temporales en la sesión
            request.session['verification_code'] = verification_code
            
            # Convertir cleaned_data a un formato serializable
            form_data = {}
            for key, value in user_creation_form.cleaned_data.items():
                if hasattr(value, 'read'):  # Es un archivo
                    # No almacenar archivos en la sesión
                    continue
                else:
                    form_data[key] = value
            
            request.session['user_form_data'] = form_data
            
            # Almacenar archivos temporalmente en el sistema de archivos
            import tempfile
            import os
            temp_files = {}
            for key, file in request.FILES.items():
                if file:
                    # Crear archivo temporal
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=f'_{file.name}')
                    for chunk in file.chunks():
                        temp_file.write(chunk)
                    temp_file.close()
                    temp_files[key] = {
                        'path': temp_file.name,
                        'name': file.name,
                        'content_type': file.content_type,
                        'size': file.size
                    }
            
            request.session['temp_files'] = temp_files

            # Enviar email
            email_text = 'Bienvenido al Centro Fray Bartolome de las Casas, para completar su registro ingrese el siguiente codigo : ' + verification_code
            try:
                send_mail(
                    'Código de Verificación - Centro Fray Bartolome de las Casas',
                    email_text,
                    settings.DEFAULT_FROM_EMAIL,
                    [user_creation_form.cleaned_data['email']],
                    fail_silently=False,
                )
                # Redirigir a la página de verificación
                return redirect('principal:verify_email')
            except Exception as e:
                print(f"Error al enviar email: {str(e)}")
                messages.error(request, 'Error al enviar el código de verificación. Por favor, intente nuevamente más tarde.')
        else:
            # Mostrar solo errores específicos como mensajes, excepto email y carnet
            for field, errors in user_creation_form.errors.items():
                for error in errors:
                    # No mostrar errores de email y carnet como mensajes
                    if field not in ['email', 'carnet']:
                        if field == 'password2' and 'password_mismatch' in error:
                            messages.error(request, 'Las contraseñas no coinciden. Por favor, asegúrese de escribir la misma contraseña en ambos campos.')
                        else:
                            # No mostrar errores como mensajes para que aparezcan solo en los campos
                            pass
            print(f"Errores en el formulario: {user_creation_form.errors}")
    else:
        user_creation_form = CustomUserCreationForm()

    data = {
        'form': user_creation_form
    }
    return render(request, 'registration/registro.html', data)

def verify_email(request):
    if request.method == 'POST':
        code = request.POST.get('code')
        if code == request.session.get('verification_code'):
            user_form_data = request.session.get('user_form_data')
            temp_files_data = request.session.get('temp_files', {})
            
            # Recrear archivos desde archivos temporales
            from django.core.files.uploadedfile import SimpleUploadedFile
            import os
            
            files_dict = {}
            for key, file_info in temp_files_data.items():
                if os.path.exists(file_info['path']):
                    with open(file_info['path'], 'rb') as f:
                        files_dict[key] = SimpleUploadedFile(
                            file_info['name'],
                            f.read(),
                            content_type=file_info['content_type']
                        )
            
            user_creation_form = CustomUserCreationForm(data=user_form_data, files=files_dict)
            if user_creation_form.is_valid():
                user = user_creation_form.save(commit=True)
                messages.success(request, f"Usuario {user.username} creado correctamente")
                
                # Limpiar archivos temporales
                for file_info in temp_files_data.values():
                    if os.path.exists(file_info['path']):
                        os.unlink(file_info['path'])
                
                # Limpiar sesión
                del request.session['verification_code']
                del request.session['user_form_data']
                if 'temp_files' in request.session:
                    del request.session['temp_files']

                # Enviar correo de confirmación de registro
                confirmation_subject = 'Registro Exitoso - Centro Fray Bartolome de las Casas'
                confirmation_message = f'Usted se ha registrado satisfactoriamente. Su Nombre de Usuario es: {user.username}'
                try:
                    send_mail(
                        confirmation_subject,
                        confirmation_message,
                        settings.DEFAULT_FROM_EMAIL,
                        [user.email],
                        fail_silently=False,
                    )
                except Exception as e:
                    print(f"Error al enviar correo de confirmación: {str(e)}")
                    # Considerar si se debe mostrar un mensaje de error al usuario o simplemente loguear

                return redirect('login')
            else:
                # Limpiar archivos temporales en caso de error
                for file_info in temp_files_data.values():
                    if os.path.exists(file_info['path']):
                        os.unlink(file_info['path'])
                error_message = 'Error al crear el usuario. Por favor, intente nuevamente.'
        else:
            error_message = 'Código incorrecto. Por favor, intente nuevamente.'

        return render(request, 'registration/verify_email.html', {'error': error_message})

    return render(request, 'registration/verify_email.html')

# Vista para manejar la redirección después del login

class LoginRedirectView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            # Verificar si el usuario fue creado automáticamente desde datos archivados
            if request.session.get('usuario_creado_automaticamente'):
                fuente = request.session.get('usuario_creado_desde', 'datos_archivados')
                
                if fuente == 'datos_dinamicos':
                    messages.success(
                        request, 
                        f'¡Bienvenido de vuelta, {user.get_full_name() or user.username}! '
                        'Su cuenta ha sido creada automáticamente desde los datos archivados. '
                        'Ahora puede acceder a todos los servicios del sistema. '
                        'Se le ha enviado un email con los detalles de su cuenta.'
                    )
                else:
                    messages.success(
                        request, 
                        f'¡Bienvenido de vuelta, {user.get_full_name() or user.username}! '
                        'Su cuenta ha sido reactivada automáticamente desde los datos archivados. '
                        'Ahora puede acceder a todos los servicios del sistema. '
                        'Se le ha enviado un email con los detalles de su cuenta.'
                    )
                
                # Limpiar las variables de sesión
                del request.session['usuario_creado_automaticamente']
                if 'usuario_creado_desde' in request.session:
                    del request.session['usuario_creado_desde']
            
            # Redirección normal según el grupo del usuario
            if user.groups.filter(name='Profesores').exists() or user.groups.filter(name='Administracion').exists() or user.groups.filter(name='Secretaria').exists():
                return redirect('principal:profile')  # Redirige a la página de perfil del profesor o el admin
            else:
                return redirect('principal:cursos')  # Redirige a la página de cursos para otros usuarios
        return redirect('principal:home') # Redirige a home si no está autenticado (aunque LoginRequiredMixin ya lo manejaría)



    # Pagina de Perfil


class ProfileView(BaseContextMixin, TemplateView):
    template_name = 'profile/profile.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Obtener el primer grupo del usuario de forma segura
        user_group = user.groups.first()
        group_name = user_group.name if user_group else None
        
        # Asegurar que group_name esté en el contexto (BaseContextMixin ya lo hace, pero por seguridad)
        context['group_name'] = group_name

        if group_name == 'Profesores':
            # Obtener todos los cursos asignados al profesor del curso académico activo
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            if curso_academico_activo:
                assigned_courses = Curso.objects.filter(teacher=user, curso_academico=curso_academico_activo)
            else:
                assigned_courses = Curso.objects.none()
            context['assigned_courses'] = assigned_courses
            
            # Obtener las solicitudes de inscripción pendientes para los cursos del profesor
            pending_solicitudes = SolicitudInscripcion.objects.filter(
                curso__teacher=user,
                estado='pendiente'
            ).order_by('-fecha_solicitud')
            context['pending_solicitudes'] = pending_solicitudes
        elif group_name == 'Estudiantes':
            # Obtener los cursos en los que el estudiante está inscrito y que pertenecen al curso académico activo
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            if curso_academico_activo:
                enrolled_courses = Curso.objects.filter(matriculas__student=user, curso_academico=curso_academico_activo)
                
                # Separar cursos por estado de solicitud
                approved_courses = []
                pending_courses = []
                
                # Para cada curso inscrito, obtener información adicional sobre solicitudes
                for course in enrolled_courses:
                    # Verificar si hay una solicitud de inscripción para este curso
                    try:
                        solicitud = SolicitudInscripcion.objects.get(
                            estudiante=user,
                            curso=course
                        )
                        course.solicitud_estado = solicitud.estado
                        course.fecha_revision = solicitud.fecha_revision
                        course.revisado_por = solicitud.revisado_por
                        
                        # Separar por estado solo para cursos en inscripción
                        if course.status in ['I', 'IT'] and solicitud.estado == 'pendiente':
                            pending_courses.append(course)
                        else:
                            approved_courses.append(course)
                    except SolicitudInscripcion.DoesNotExist:
                        course.solicitud_estado = None
                        course.fecha_revision = None
                        course.revisado_por = None
                        approved_courses.append(course)
                
                context['enrolled_courses'] = approved_courses
                context['pending_courses'] = pending_courses
            else:
                context['enrolled_courses'] = Curso.objects.none()
                context['pending_courses'] = Curso.objects.none()
        elif group_name in ['Administracion', 'Secretaria']:
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            if curso_academico_activo:
                all_courses = Curso.objects.filter(curso_academico=curso_academico_activo)
            else:
                all_courses = Curso.objects.none()
            context['all_courses'] = all_courses
        
        return context

# Vista de los Cursos


class MatriculasListView(BaseContextMixin, ListView):
    model = Matriculas
    template_name = 'matriculas_list.html'
    context_object_name = 'matriculas'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtering by CursoAcademico
        curso_academico_id = self.request.GET.get('curso_academico')
        if curso_academico_id:
            queryset = queryset.filter(curso_academico__id=curso_academico_id)

        # Filtering by Curso
        curso_id = self.request.GET.get('curso')
        if curso_id:
            queryset = queryset.filter(course__id=curso_id)

        # Filtering by Estudiante
        student_id = self.request.GET.get('student')
        if student_id:
            queryset = queryset.filter(student__id=student_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cursos_academicos'] = CursoAcademico.objects.all()
        context['cursos'] = Curso.objects.all()
        context['estudiantes'] = User.objects.filter(groups__name='Estudiantes')
        return context


# Vistas para Calificaciones
class CalificacionesListView(BaseContextMixin, ListView):
    model = Calificaciones
    template_name = 'calificaciones_list.html'
    context_object_name = 'calificaciones'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtering by CursoAcademico
        curso_academico_id = self.request.GET.get('curso_academico')
        if curso_academico_id:
            queryset = queryset.filter(curso_academico__id=curso_academico_id)

        # Filtering by Curso
        curso_id = self.request.GET.get('curso')
        if curso_id:
            queryset = queryset.filter(course__id=curso_id)

        # Filtering by Estudiante
        student_id = self.request.GET.get('student')
        if student_id:
            queryset = queryset.filter(student__id=student_id)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cursos_academicos'] = CursoAcademico.objects.all()
        context['cursos'] = Curso.objects.all()
        context['estudiantes'] = User.objects.filter(groups__name='Estudiantes')
        return context


class StudentCourseAttendanceView(BaseContextMixin, ListView):
    model = Asistencia
    template_name = 'student_asistencias.html'
    context_object_name = 'asistencias'

    def get_queryset(self):
        queryset = super().get_queryset()
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']
        queryset = queryset.filter(student__id=student_id, course__id=course_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']
        context['student'] = User.objects.get(id=student_id)
        context['course'] = Curso.objects.get(id=course_id)
        context['student'] = User.objects.get(id=student_id)
        return context


# Vistas para Asistencias
class AsistenciasListView(ListView):
    model = Asistencia
    template_name = 'asistencias_list.html'
    context_object_name = 'asistencias'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtrar por Curso Académico
        curso_academico_id = self.request.GET.get('curso_academico')
        if curso_academico_id:
            queryset = queryset.filter(course__curso_academico__id=curso_academico_id)

        # Filtrar por Curso
        curso_id = self.request.GET.get('curso')
        if curso_id:
            queryset = queryset.filter(course__id=curso_id)

        # Filtrar por Estudiante
        estudiante_id = self.request.GET.get('estudiante')
        if estudiante_id:
            queryset = queryset.filter(student__id=estudiante_id)

        # Filtrar por Fecha
        fecha_asistencia = self.request.GET.get('fecha')
        if fecha_asistencia:
            queryset = queryset.filter(date=fecha_asistencia)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cursos_academicos'] = CursoAcademico.objects.all()
        context['cursos'] = Curso.objects.all()
        context['estudiantes'] = User.objects.filter(groups__name='Estudiantes')
        context['selected_curso_academico'] = self.request.GET.get('curso_academico')
        
        # Calcular porcentaje de asistencia cuando se filtra por curso y estudiante
        curso_id = self.request.GET.get('curso')
        estudiante_id = self.request.GET.get('estudiante')
        
        if curso_id and estudiante_id:
            # Obtener todas las asistencias del estudiante en el curso
            total_asistencias = Asistencia.objects.filter(course_id=curso_id, student_id=estudiante_id).count()
            presentes = Asistencia.objects.filter(course_id=curso_id, student_id=estudiante_id, presente=True).count()
            
            if total_asistencias > 0:
                porcentaje = (presentes / total_asistencias) * 100
                context['porcentaje_asistencia'] = round(porcentaje, 2)
                context['total_asistencias'] = total_asistencias
                context['presentes'] = presentes
                context['ausentes'] = total_asistencias - presentes
        context['selected_curso'] = self.request.GET.get('curso')
        context['selected_estudiante'] = self.request.GET.get('estudiante')
        context['selected_fecha'] = self.request.GET.get('fecha')
        return context


class StudentCourseNotesView(BaseContextMixin, ListView):
    model = Calificaciones
    template_name = 'student_notes.html'  # New template for student notes
    context_object_name = 'calificaciones'

    def get_queryset(self):
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']
        # Filter grades for the specific student and course
        return Calificaciones.objects.filter(student__id=student_id, course__id=course_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student_id = self.kwargs['student_id']
        course_id = self.kwargs['course_id']

        student = User.objects.get(id=student_id)
        course = Curso.objects.get(id=course_id)

        context['student'] = student
        context['course'] = course

        print(f"[DEBUG] StudentCourseNotesView - student_id: {student_id}, course_id: {course_id}")

        # Get all Calificaciones for the student and course
        calificaciones_for_student_course = Calificaciones.objects.filter(
            student=student,
            course=course
        )
        print(f"[DEBUG] Calificaciones found: {calificaciones_for_student_course.count()}")

        all_notes = []
        total_score = 0
        num_grades = 0

        for calificacion in calificaciones_for_student_course:
            print(f"[DEBUG] Processing Calificacion ID: {calificacion.id}")
            for nota in calificacion.notas.all():
                print(f"[DEBUG] Adding Nota: {nota.valor} (ID: {nota.id})")
                all_notes.append(nota)
                if nota.valor is not None:
                    total_score += nota.valor
                    num_grades += 1

        if num_grades > 0:
            average_score = total_score / num_grades
        else:
            average_score = 0

        context['all_notes'] = all_notes
        context['average_score'] = average_score
        return context


class CoursesView(BaseContextMixin, TemplateView):
    template_name = 'cursos.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        if curso_academico_activo:
            courses = Curso.objects.filter(curso_academico=curso_academico_activo)
        else:
            courses = Curso.objects.none()
        student = self.request.user if self.request.user.is_authenticated else None

        for item in courses:
            if student:
                registration = Matriculas.objects.filter(
                    course=item, student=student).first()
                item.is_enrolled = registration is not None
            else:
                item.is_enrolled = False

            # Calcular el conteo de inscripciones para todos los cursos
            enrollment_count = Matriculas.objects.filter(course=item).count()
            item.enrollment_count = enrollment_count

        context['courses'] = courses
        # Asegurarse de que group_name esté en el contexto
        user = self.request.user
        if user.is_authenticated:
            group = Group.objects.filter(user=user).first()
            if group:
                context['group_name'] = group.name
            else:
                context['group_name'] = None
        else:
            context['group_name'] = None

        return context

# Crear nuevo Curso


class CourseCreateView(LoginRequiredMixin, CreateView):
    model = Curso
    form_class = CourseForm
    template_name = 'create_course.html'
    success_url = reverse_lazy('principal:cursos')

    def form_valid(self, form):
        # Asigna el curso académico activo al curso
        active_academic_course = CursoAcademico.objects.filter(activo=True).first()
        if active_academic_course:
            form.instance.curso_academico = active_academic_course
        messages.success(self.request, 'El Curso se guardo correctamente')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(
            self.request, 'Ha ocurrido un ERROR al guardar el Curso')
        return self.render_to_response(self.get_context_data(form=form))

# Vista para inscribirse a un curso


@login_required
def inscribirse_curso(request, curso_id):
    # Obtener el curso
    curso = Curso.objects.get(id=curso_id)
    estudiante = request.user

    # Verificar si ya está inscrito
    inscripcion_existente = Matriculas.objects.filter(
        course=curso, student=estudiante).exists()

    if not inscripcion_existente:
        # Obtener el curso académico activo
        curso_academico = CursoAcademico.objects.filter(activo=True).first()
        
        if not curso_academico:
            messages.error(request, 'No hay un curso académico activo configurado. Contacte al administrador.')
            return redirect('principal:cursos')
            
        # Crear nueva matrícula asignada al curso académico activo
        matricula = Matriculas(
            course=curso, 
            student=estudiante, 
            activo=True,
            curso_academico=curso_academico,
            estado='P'  # Estado inicial: Pendiente
        )
        matricula.save()
        messages.success(
            request, f'Te has inscrito exitosamente al curso {curso.name} para el año académico {curso_academico.nombre}')
    else:
        messages.info(request, 'Ya estás inscrito en este curso')

    return redirect('principal:cursos')

# Vista para editar un curso


class CourseUpdateView(BaseContextMixin, UpdateView):
    model = Curso
    form_class = CourseForm
    template_name = 'create_course.html'  # Reutilizamos el mismo template
    success_url = reverse_lazy('principal:cursos')

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        # Verificar si el curso tiene un formulario de aplicación
        try:
            formulario = FormularioAplicacion.objects.filter(curso=obj).first()
            if formulario:
                obj.formulario_aplicacion = formulario
        except Exception:
            pass
        return obj

    @override
    def form_valid(self, form):
        messages.success(self.request, 'El Curso se actualizó correctamente')
        return super().form_valid(form)

    @override
    def form_invalid(self, form):
        messages.error(
            self.request, 'Ha ocurrido un ERROR al actualizar el Curso')
        return self.render_to_response(self.get_context_data(form=form))


# Vista para eliminar un curso
@login_required
def eliminar_curso(request, curso_id):
    # Verificar si el usuario pertenece al grupo 'Secretaria'
    if request.user.groups.filter(name='Secretaria').exists():
        try:
            # Obtener el curso
            curso = Curso.objects.get(id=curso_id)
            nombre_curso = curso.name

            # Eliminar el curso
            curso.delete()
            messages.success(
                request, f'El curso {nombre_curso} ha sido eliminado correctamente')
        except Curso.DoesNotExist:
            messages.error(request, 'El curso no existe')
    else:
        messages.error(request, 'No tienes permisos para eliminar cursos')

    return redirect('principal:cursos')

# Mostrar lista de alumnos y notas a los profesores


class StudentListNotasView(BaseContextMixin, ListView):
    model = Matriculas
    template_name = 'student_list_notas.html'
    context_object_name = 'matriculas'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'student',
            'course',
            'course__teacher',
            'calificaciones'
        )

        # Verificar si se está accediendo desde la URL con course_id
        course_id = self.kwargs.get('course_id')
        if course_id:
            queryset = queryset.filter(course__id=course_id)
            return queryset

        # Si no hay course_id en la URL, usar los filtros normales
        search_query = self.request.GET.get('search_query')
        course_filter = self.request.GET.get('course')
        teacher_filter = self.request.GET.get('teacher')

        if search_query:
            queryset = queryset.filter(
                Q(student__username__icontains=search_query) |
                Q(student__first_name__icontains=search_query) |
                Q(student__last_name__icontains=search_query)
            )


        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        course_id = self.kwargs.get('course_id')
        if course_id:
            course = get_object_or_404(Curso, id=course_id)
            context['course'] = course

            # Obtener el curso académico activo
            curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
            context['curso_academico'] = curso_academico_activo

            # Obtener solo las matrículas activas para este curso y curso académico activo
            active_enrollments = Matriculas.objects.filter(
                course=course,
                activo=True,
                curso_academico=curso_academico_activo
            )

            student_data = []
            for enrollment in active_enrollments:
                student = enrollment.student
                # Buscar calificación por curso, estudiante y curso académico
                calificacion = Calificaciones.objects.filter(
                    course=course,
                    student=student,
                    curso_academico=curso_academico_activo
                ).first()

                all_notes = []
                if calificacion:
                    # Obtener todas las notas individuales relacionadas con esta calificación
                    all_notes = list(calificacion.notas.all().order_by('fecha_creacion'))

                student_data.append({
                    'calificacion_id': calificacion.id if calificacion else None,
                    'name': student.get_full_name(),
                    'notas': all_notes, # Lista de notas individuales
                    'average': calificacion.average if calificacion else None,
                    'matricula_id': enrollment.id,
                    'student_id': student.id, # Add student.id here
                })
            context['student_data'] = student_data
        else:
            context['courses'] = CursoAcademico.objects.all()
            context['teachers'] = User.objects.filter(groups__name='Docente')
        
        # Obtener el curso académico activo
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        
        # Obtener solo las matrículas activas para este curso y curso académico activo
        active_enrollments = Matriculas.objects.filter(
            course=course, 
            activo=True,
            curso_academico=curso_academico_activo
        )
        
        student_data = []
        for enrollment in active_enrollments:
            student = enrollment.student
            # Buscar calificación por curso, estudiante y curso académico
            calificacion = Calificaciones.objects.filter(
                course=course, 
                student=student,
                curso_academico=curso_academico_activo
            ).first()
            
            all_notes = []
            notas_list = []
            if calificacion:
                # Obtener todas las notas individuales relacionadas con esta calificación
                notas_list = list(calificacion.notas.all().order_by('fecha_creacion').values_list('valor', flat=True))

            student_data.append({
                'calificacion_id': calificacion.id if calificacion else None,
                'name': student.get_full_name(),
                'notas': notas_list, # Lista de notas individuales
                'average': calificacion.average if calificacion else None,
                'matricula_id': enrollment.id,
                'student_id': student.id, # Add student.id here
            })

        context['course'] = course
        context['student_data'] = student_data
        context['curso_academico'] = curso_academico_activo
        return context
# Agregar Notas de los estudiantes

class AddNotaView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):

        return super().dispatch(request, *args, **kwargs)

    def get(self, request, matricula_id):
        matricula = get_object_or_404(Matriculas, id=matricula_id)
        print(f"[GET] Matricula ID: {matricula_id}")
        print(f"[GET] Matricula Course ID: {matricula.course.id}")
        print(f"[GET] Matricula Student ID: {matricula.student.id}")
        print(f"[GET] Matricula Curso Academico ID: {matricula.curso_academico.id if matricula.curso_academico else 'None'}")

        try:
            # Buscar calificación por curso, estudiante y curso académico de la matrícula
            calificacion = Calificaciones.objects.get(
                course=matricula.course, 
                student=matricula.student,
                curso_academico=matricula.curso_academico
            )
            print("[GET] Existing Calificacion found.")
            form = CalificacionesForm(instance=calificacion)
            formset = NotaIndividualFormSet(instance=calificacion) # Instancia el formset con la calificación existente
        except Calificaciones.DoesNotExist:
            print("[GET] Calificaciones.DoesNotExist raised. No existing Calificacion found.")
            form = CalificacionesForm(initial={
                'matricula': matricula,
                'course': matricula.course,
                'student': matricula.student,
                'curso_academico': matricula.curso_academico
            }) # Pasa los valores iniciales para los campos de Calificaciones
            formset = NotaIndividualFormSet() # Instancia un formset vacío
        
        context = {
            'form': form,
            'formset': formset, # Añade el formset al contexto
            'matricula': matricula
        }
        return render(request, 'add_nota.html', context)

    def post(self, request, matricula_id):
        matricula = get_object_or_404(Matriculas, id=matricula_id)
        print(f"[POST] Matricula ID: {matricula_id}")
        print(f"[POST] Matricula Course ID: {matricula.course.id}")
        print(f"[POST] Matricula Student ID: {matricula.student.id}")
        print(f"[POST] Matricula Curso Academico ID: {matricula.curso_academico.id if matricula.curso_academico else 'None'}")

        try:
            calificacion = Calificaciones.objects.get(
                course=matricula.course,
                student=matricula.student,
                curso_academico=matricula.curso_academico
            )
            print("[POST] Existing Calificacion found.")
            form = CalificacionesForm(request.POST, instance=calificacion)
        except Calificaciones.DoesNotExist:
            print("[POST] Calificaciones.DoesNotExist raised. No existing Calificacion found. Creating new one.")
            form = CalificacionesForm(request.POST)

        if form.is_valid():
            calificacion = form.save(commit=False)
            calificacion.matricula = matricula
            calificacion.course = matricula.course
            calificacion.student = matricula.student
            calificacion.curso_academico = matricula.curso_academico
            calificacion.save() # <-- Guarda la instancia de Calificaciones aquí

            # Ahora que calificacion tiene un PK, podemos inicializar el formset
            formset = NotaIndividualFormSet(request.POST, instance=calificacion)

            if formset.is_valid():
                
                formset.save()
                messages.success(request, 'Notas guardadas correctamente.')
                return redirect('principal:student_list_notas_by_course', course_id=matricula.course.id)
            else:
                print(f"[POST] Formset is NOT valid. Errors: {formset.errors}")
                messages.error(request, 'Error al guardar las notas individuales.')
        else:
            print(f"[POST] CalificacionesForm errors: {form.errors}")
            messages.error(request, 'Error al guardar la calificación principal.')

        context = {
            'form': form,
            'formset': formset, # Asegúrate de pasar el formset al contexto incluso si hay errores
            'matricula': matricula
        }
        return render(request, 'add_nota.html', context)

#esto es de la ia
""" def crear_cursos(request):
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES) # Asegúrate de incluir request.FILES aquí
        if form.is_valid():
            form.save()
            # Por ejemplo, puedes añadir un mensaje de éxito y redirigir
            # messages.success(request, 'Curso creado exitosamente!')
            # return redirect('nombre_de_tu_url_de_cursos')
    else:
        form = CourseForm()
    return render(request, 'create_course.html', {'form': form})


def editar_curso(request, course_id):
    course = get_object_or_404(Curso, id=course_id)
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES, instance=course) # Asegúrate de incluir request.FILES aquí
        if form.is_valid():
            form.save()
            # Por ejemplo, puedes añadir un mensaje de éxito y redirigir
            # messages.success(request, 'Curso actualizado exitosamente!')
            # return redirect('nombre_de_tu_url_de_cursos')
    else:
        form = CourseForm(instance=course)
    return render(request, 'edit_course.html', {'form': form, 'course': course})

 """

#vistas para historicos

def historico_alumno(request, student_id):
        matriculas = Matriculas.objects.filter(student_id=student_id).select_related('curso_academico')
        return render(request, 'historico.html', {'matriculas': matriculas})


# Agregando asistencias

class AsistenciaView(BaseContextMixin, TemplateView):
    template_name = 'asistencias.html'
    
    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course_id = kwargs['course_id']
        course = get_object_or_404(Curso, id=course_id)
        
        # Obtener el curso académico activo
        curso_academico_activo = CursoAcademico.objects.filter(activo=True).first()
        
        # Filtrar asistencias por curso y ordenar por fecha descendente
        asistencias = Asistencia.objects.filter(
            course=course,
            student__matriculas__curso_academico=curso_academico_activo,
            student__matriculas__activo=True
        ).select_related('student', 'course').order_by('-date')
        
        # Obtener todas las matrículas activas para este curso en el curso académico activo
        matriculas = Matriculas.objects.filter(
            course=course,
            activo=True,
            curso_academico=curso_academico_activo
        ).select_related('student')  # Optimizar consulta de estudiantes
        
        # Filtrar por fecha si se proporciona en la solicitud
        fecha_filtro = self.request.GET.get('fecha')
        if fecha_filtro:
            asistencias = asistencias.filter(date=fecha_filtro)
        
        # Calcular la cantidad de asistencias registradas (fechas únicas)
        asistencias_registradas = Asistencia.objects.filter(course=course).values('date').distinct().count()
        
        # Obtener la cantidad total de clases del curso
        cantidad_total_clases = course.class_quantity
        
        # Calcular las clases restantes
        clases_restantes = cantidad_total_clases - asistencias_registradas

        context['course'] = course
        context['asistencias'] = asistencias
        context['matriculas'] = matriculas
        context['curso_academico'] = curso_academico_activo
        context['cantidad_total_clases'] = cantidad_total_clases
        context['asistencias_registradas'] = asistencias_registradas
        context['clases_restantes'] = clases_restantes
        return context


class AddAsistenciaView(LoginRequiredMixin, TemplateView):
    template_name='add_asistencias.html'

    @override
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        course_id = kwargs['course_id']
        course = Curso.objects.get(id=course_id)
        matriculas = Matriculas.objects.filter(course=course)
        # Obtener las asistencias existentes para este curso
        asistencias = Asistencia.objects.order_by('date')
        context['course'] = course
        context['matriculas'] = matriculas
        context['asistencias'] = asistencias
        context['today'] = date.today()
        return context

    def post(self, request, course_id):
        course = Curso.objects.get(id=course_id)
        matriculas = Matriculas.objects.filter(course=course)

        if request.method == 'POST':
            date = request.POST.get('date')

            for matricula in matriculas:
                absent = request.POST.get('asistencia_' + str(matricula.id))
                # Buscar si ya existe un registro de asistencia para este estudiante en esta fecha
                asistencia, created = Asistencia.objects.get_or_create(
                    student=matricula.student,
                    course=course,
                    date=date,
                    defaults={'presente': not bool(absent)}
                )
                # Si el registro ya existía, actualizar el estado de presente
                if not created:
                    asistencia.presente = not bool(absent)
                    asistencia.save()
        # Redirigir a la misma página para mostrar las asistencias actualizadas
        return redirect('principal:asistencias', course_id=course_id)


# Eliminar asistencia
def eliminar_asistencia(request, asistencia_id):
    # Obtener la asistencia o devolver 404 si no existe
    asistencia = get_object_or_404(Asistencia, id=asistencia_id)
    
    # Guardar el ID del curso antes de eliminar la asistencia
    course_id = asistencia.course.id
    
    # Eliminar la asistencia
    asistencia.delete()
    
    # Redirigir a la página de asistencias del curso
    return redirect('principal:asistencias', course_id=course_id)


def add_asistencias(request, course_id):
    course = get_object_or_404(Curso, id=course_id)
    matriculas = Matriculas.objects.filter(course=course, activo=True)

    if request.method == 'POST':
        date_str = request.POST.get('date')
        try:
            attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Formato de fecha inválido.")
            return redirect('principal:add_asistencias', course_id=course.id)

        # Eliminar asistencias existentes para esta fecha y curso
        Asistencia.objects.filter(course=course, date=attendance_date).delete()

        for matricula in matriculas:
            is_absent = request.POST.get(f'asistencia_{matricula.id}')
            presente = not bool(is_absent) # Si está marcado, significa que está ausente, por lo tanto, no presente

            Asistencia.objects.create(
                course=course,
                student=matricula.student,
                date=attendance_date,
                presente=presente
            )
        messages.success(request, "Asistencias guardadas correctamente.")
        return redirect('principal:asistencias', course_id=course.id) # Redirige a la página de asistencias del curso
    
    today = date.today()
    context = {
        'course': course,
        'matriculas': matriculas,
        'today': today,
    }
    return render(request, 'add_asistencias.html', context)


@login_required
def undo_last_asistencia(request, course_id):
    course = get_object_or_404(Curso, id=course_id)

    # Encontrar la fecha más reciente para la que se registró asistencia en este curso
    latest_attendance_date_obj = Asistencia.objects.filter(course=course).aggregate(Max('date'))
    latest_attendance_date = latest_attendance_date_obj['date__max']

    if latest_attendance_date:
        # Eliminar todos los registros de asistencia para este curso en la fecha más reciente
        Asistencia.objects.filter(course=course, date=latest_attendance_date).delete()
        messages.success(request, f"La asistencia del {latest_attendance_date.strftime('%d-%m-%Y')} ha sido deshecha correctamente.")
    else:
        messages.info(request, "No hay asistencias registradas para deshacer en este curso.")

    return redirect('principal:asistencias', course_id=course.id)

    

# Vistas para el sistema de formularios de aplicación a cursos

class SecretariaRequiredMixin(UserPassesTestMixin):
    """
    Mixin que verifica que el usuario pertenezca al grupo Secretaria.
    """
    def test_func(self):
        return self.request.user.groups.filter(name='Secretaria').exists()

class ProfesorRequiredMixin(UserPassesTestMixin):
    """
    Mixin que verifica que el usuario pertenezca al grupo Profesores.
    """
    def test_func(self):
        return self.request.user.groups.filter(name='Profesores').exists()

class FormularioAplicacionListView(LoginRequiredMixin, SecretariaRequiredMixin, ListView):
    """
    Vista para listar los formularios de aplicación creados por el grupo secretaria.
    """
    model = FormularioAplicacion
    template_name = 'formularios/formulario_list.html'
    context_object_name = 'formularios'

    def get_queryset(self):
        return FormularioAplicacion.objects.all().order_by('-fecha_modificacion')

class FormularioAplicacionCreateView(LoginRequiredMixin, SecretariaRequiredMixin, CreateView):
    """
    Vista para crear un nuevo formulario de aplicación.
    """
    model = FormularioAplicacion
    form_class = FormularioAplicacionForm
    template_name = 'formularios/formulario_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Formulario de Aplicación'
        
        # Verificar si se recibió un curso_id en la URL
        curso_id = self.request.GET.get('curso_id')
        if curso_id:
            try:
                curso = Curso.objects.get(id=curso_id)
                context['curso_preseleccionado'] = curso
                return context
            except Curso.DoesNotExist:
                pass
        
        # Obtener los cursos que no tienen formulario de aplicación
        cursos_sin_formulario = Curso.objects.filter(
            curso_academico__activo=True
        ).exclude(
            id__in=FormularioAplicacion.objects.values_list('curso_id', flat=True)
        )
        context['cursos'] = cursos_sin_formulario
        
        return context
        
    def get_initial(self):
        """
        Establece valores iniciales para el formulario.
        """
        initial = super().get_initial()
        initial['descripcion'] = "Por favor, conteste responsablemente todas las preguntas que le hacemos a continuación, eso ayudará a los profesores a una mejor organización del curso. Muchas gracias."
        return initial
    
    def form_valid(self, form):
        curso_id = self.request.POST.get('curso')
        if curso_id:
            curso = get_object_or_404(Curso, id=curso_id)
            form.instance.curso = curso
            response = super().form_valid(form)
            
            # Imprimir información de depuración
            print(f"DEBUG: Formulario creado para curso {curso.name} (ID: {curso.id})")
            print(f"DEBUG: ID del formulario: {self.object.id}")
            print(f"DEBUG: Verificando relación: {FormularioAplicacion.objects.filter(curso=curso).exists()}")
            
            # Limpiar la caché de la sesión para forzar una recarga de los datos
            if 'cursos_con_formularios' in self.request.session:
                del self.request.session['cursos_con_formularios']
            
            return response
        else:
            messages.error(self.request, 'Debe seleccionar un curso.')
            return self.form_invalid(form)
    
    def get_success_url(self):
        # Añadir parámetro para indicar que viene de la creación del formulario
        return reverse('principal:formulario_preguntas', kwargs={'pk': self.object.pk}) + '?from_create=1'

class FormularioAplicacionUpdateView(LoginRequiredMixin, SecretariaRequiredMixin, UpdateView):
    """
    Vista para editar un formulario de aplicación existente.
    """
    model = FormularioAplicacion
    form_class = FormularioAplicacionForm
    template_name = 'formularios/formulario_form.html'
    
    def get_object(self, queryset=None):
        """
        Obtiene el objeto que se va a editar y maneja posibles errores.
        """
        try:
            obj = super().get_object(queryset)
            return obj
        except Exception as e:
            # Registrar el error para depuración
            print(f"Error al obtener el objeto FormularioAplicacion: {e}")
            # Redirigir a la lista de formularios con un mensaje de error
            messages.error(self.request, f"No se pudo encontrar el formulario solicitado. Error: {e}")
            return None
    
    def get(self, request, *args, **kwargs):
        """
        Maneja la solicitud GET y redirige si no se encuentra el objeto.
        """
        self.object = self.get_object()
        if self.object is None:
            return redirect('principal:formulario_list')
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Editar Formulario de Aplicación'
        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Formulario de aplicación actualizado correctamente.')
        return response
    
    def get_success_url(self):
        return reverse('principal:formulario_preguntas', kwargs={'pk': self.object.pk})

class FormularioPreguntasView(LoginRequiredMixin, SecretariaRequiredMixin, UpdateView):
    """
    Vista para gestionar las preguntas de un formulario de aplicación.
    """
    model = FormularioAplicacion
    template_name = 'formularios/formulario_preguntas.html'
    fields = []  # No necesitamos campos para esta vista
    
    def get(self, request, *args, **kwargs):
        # Limpiar todos los mensajes existentes para evitar duplicados
        storage = messages.get_messages(request)
        # Consumir todos los mensajes para limpiarlos
        for _ in storage:
            pass
        
        # Si viene de la creación del formulario, mostrar un solo mensaje
        if request.GET.get('from_create'):
            messages.success(request, 'Formulario de aplicación creado correctamente.')
        
        response = super().get(request, *args, **kwargs)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        formulario = self.get_object()
        
        if self.request.POST:
            context['pregunta_formset'] = PreguntaFormularioFormSet(
                self.request.POST, instance=formulario
            )
        else:
            context['pregunta_formset'] = PreguntaFormularioFormSet(instance=formulario)
        
        context['formulario'] = formulario
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        pregunta_formset = context['pregunta_formset']
        
        # Imprimir datos del formulario para depuración
        print("\n=== DATOS DEL FORMULARIO RECIBIDOS ===")
        for key, value in self.request.POST.items():
            print(f"{key}: '{value}'")
        print("=====================================\n")
        
        if pregunta_formset.is_valid():
            # Guardar las preguntas
            preguntas = pregunta_formset.save(commit=True)
            print(f"Preguntas guardadas: {preguntas}")
            
            # Asegurarse de que el curso tenga el atributo tiene_formulario
            formulario = self.object
            curso = formulario.curso
            print(f"DEBUG: Formulario {formulario.id} asociado al curso {curso.name} (ID: {curso.id})")
            print(f"DEBUG: Verificando relación después de guardar preguntas: {FormularioAplicacion.objects.filter(curso=curso).exists()}")
            
            # Limpiar la caché de la sesión para forzar una recarga de los datos
            if 'cursos_con_formularios' in self.request.session:
                del self.request.session['cursos_con_formularios']
            
            # Si se está redirigiendo a las opciones, buscar la última pregunta creada
            if self.request.POST.get('redirect_to_options') or self.request.POST.get('save_and_continue'):
                # Obtener la última pregunta creada para este formulario
                ultima_pregunta = self.object.preguntas.order_by('-id').first()
                if ultima_pregunta:
                    # No mostrar mensaje aquí, lo mostraremos en la vista de opciones
                    # Imprimir información de depuración
                    print(f"Redirigiendo a opciones de pregunta {ultima_pregunta.id}: {ultima_pregunta.texto}")
                    # Redirigir directamente a la página de opciones de la pregunta con parámetro
                    return redirect(reverse('principal:pregunta_opciones', kwargs={'pk': ultima_pregunta.pk}) + '?from_redirect=1')
            
            # Limpiar todos los mensajes existentes antes de añadir uno nuevo
            storage = messages.get_messages(self.request)
            for _ in storage:
                pass  # Consumir todos los mensajes
            
            # Añadir un solo mensaje de éxito
            messages.success(self.request, 'Preguntas guardadas correctamente.')
            
            # Redirigir a la página de cursos
            return redirect(reverse('principal:cursos'))
        else:
            print("\n=== ERRORES DEL FORMSET ===")
            for i, form_errors in enumerate(pregunta_formset.errors):
                print(f"Formulario {i}: {form_errors}")
            print("=========================\n")
            return self.render_to_response(self.get_context_data(form=form))
    
    def get_success_url(self):
        # Redirigir a la página de cursos
        return reverse('principal:cursos')

class PreguntaOpcionesView(LoginRequiredMixin, SecretariaRequiredMixin, UpdateView):
    """
    Vista para gestionar las opciones de respuesta de una pregunta.
    """
    model = PreguntaFormulario
    template_name = 'formularios/pregunta_opciones.html'
    fields = []  # No necesitamos campos para esta vista
    
    def get(self, request, *args, **kwargs):
        # Limpiar todos los mensajes existentes para evitar duplicados
        storage = messages.get_messages(request)
        # Consumir todos los mensajes para limpiarlos
        for _ in storage:
            pass
        
        # Si viene de la redirección de una pregunta guardada, mostrar un mensaje
        if request.GET.get('from_redirect'):
            messages.success(request, 'Pregunta guardada correctamente. Ahora puedes agregar opciones de respuesta.')
        
        response = super().get(request, *args, **kwargs)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pregunta = self.get_object()
        
        if self.request.POST:
            context['opcion_formset'] = OpcionRespuestaFormSet(
                self.request.POST, instance=pregunta
            )
        else:
            # Crear el formset con la instancia de la pregunta
            # El formset ya está configurado para añadir una fila extra (extra=1)
            context['opcion_formset'] = OpcionRespuestaFormSet(instance=pregunta)
        
        context['pregunta'] = pregunta
        context['formulario'] = pregunta.formulario
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        pregunta = self.object
        
        # Imprimir información de depuración detallada
        print("\n=== DATOS DEL FORMULARIO RECIBIDOS ===")
        for key, value in request.POST.items():
            print(f"{key}: '{value}'")
        print("=====================================\n")
        
        # Procesar los datos del formulario manualmente
        try:
            # Obtener el número total de opciones
            total_opciones = int(request.POST.get('total_opciones', 0))
            print(f"Total de opciones: {total_opciones}")
            
            # Procesar cada opción
            for i in range(total_opciones):
                opcion_id = request.POST.get(f'opcion_id_{i}', '')
                texto = request.POST.get(f'texto_{i}', '')
                orden = request.POST.get(f'orden_{i}', i)
                eliminar = request.POST.get(f'eliminar_{i}', '') == 'on'
                
                print(f"Opción {i}: id='{opcion_id}', texto='{texto}', orden={orden}, eliminar={eliminar}")
                
                # Si la opción está marcada para eliminar, eliminarla
                if eliminar and opcion_id:
                    try:
                        opcion = OpcionRespuesta.objects.get(id=opcion_id)
                        opcion.delete()
                        print(f"  - Opción eliminada: {opcion}")
                        continue
                    except OpcionRespuesta.DoesNotExist:
                        pass
                
                # Si la opción ya existe, actualizarla
                if opcion_id:
                    try:
                        opcion = OpcionRespuesta.objects.get(id=opcion_id)
                        opcion.texto = texto
                        opcion.orden = int(orden) if orden else i
                        opcion.save()
                        print(f"  - Opción actualizada: {opcion}")
                    except OpcionRespuesta.DoesNotExist:
                        pass
                # Si es una nueva opción, crearla
                else:
                    opcion = OpcionRespuesta(
                        pregunta=pregunta,
                        texto=texto,
                        orden=int(orden) if orden else i
                    )
                    opcion.save()
                    print(f"  - Opción creada: {opcion}")
            
            # Limpiar todos los mensajes existentes antes de añadir uno nuevo
            storage = messages.get_messages(self.request)
            for _ in storage:
                pass  # Consumir todos los mensajes
            
            messages.success(self.request, 'Opciones de respuesta guardadas correctamente.')
            return redirect(self.get_success_url())
        except Exception as e:
            print(f"Error al procesar el formulario: {e}")
            messages.error(self.request, f'Error al guardar las opciones: {e}')
            return self.get(request, *args, **kwargs)
    

    
    def form_invalid(self, opcion_formset):
        # Mostrar errores específicos
        print("\n=== ERRORES DEL FORMSET ===")
        for i, form_errors in enumerate(opcion_formset.errors):
            print(f"Formulario {i}: {form_errors}")
            for field, errors in form_errors.items():
                for error in errors:
                    messages.error(self.request, f"Error en formulario {i}, campo {field}: {error}")
        print("=========================\n")
        
        # Preparar el contexto para renderizar la respuesta
        context = self.get_context_data()
        context['opcion_formset'] = opcion_formset
        return self.render_to_response(context)
    
    def get_success_url(self):
        return reverse('principal:formulario_preguntas', kwargs={'pk': self.object.formulario.pk})

# Vistas para los estudiantes

@login_required
def aplicar_curso(request, curso_id):
    """
    Vista para que un estudiante aplique a un curso mediante un formulario dinámico.
    """
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar si el curso tiene un formulario de aplicación
    try:
        # Intentar obtener el formulario directamente
        formulario = FormularioAplicacion.objects.get(curso_id=curso_id)
        if not formulario:
            messages.error(request, 'Este curso no tiene un formulario de aplicación disponible.')
            return redirect('principal:cursos')
    except FormularioAplicacion.DoesNotExist:
        messages.error(request, 'Este curso no tiene un formulario de aplicación disponible.')
        return redirect('principal:cursos')
    
    # Verificar si el estudiante ya ha aplicado a este curso
    solicitud_existente = SolicitudInscripcion.objects.filter(
        curso=curso,
        estudiante=request.user
    ).first()
    
    if solicitud_existente:
        messages.info(request, 'Ya has aplicado a este curso. Tu solicitud está en proceso de revisión.')
        return redirect('principal:cursos')
    
    # Verificar si el estudiante ya está matriculado en este curso
    matricula_existente = Matriculas.objects.filter(
        course=curso,
        student=request.user
    ).exists()
    
    if matricula_existente:
        messages.info(request, 'Ya estás matriculado en este curso.')
        return redirect('principal:cursos')
    
    # Obtener las preguntas del formulario
    preguntas = formulario.preguntas.all().order_by('orden')
    
    if request.method == 'POST':
        try:
            # Crear la solicitud de inscripción
            solicitud = SolicitudInscripcion.objects.create(
                curso=curso,
                estudiante=request.user,
                formulario=formulario,
                estado='pendiente'  # Asegurarse de que el estado sea 'pendiente'
            )
            
            # Mensaje de depuración
            print(f"DEBUG: Solicitud creada - ID: {solicitud.id}, Estado: {solicitud.estado}, Curso: {curso.name}, Estudiante: {request.user.username}")
            
            # Verificar que la solicitud se haya creado correctamente
            solicitud_verificada = SolicitudInscripcion.objects.filter(
                id=solicitud.id
            ).first()
            
            if solicitud_verificada:
                print(f"DEBUG: Solicitud verificada - ID: {solicitud_verificada.id}, Estado: {solicitud_verificada.estado}")
            else:
                print("ERROR: No se pudo verificar la solicitud creada")
        except Exception as e:
            print(f"ERROR al crear la solicitud: {str(e)}")
            raise
        
        # Procesar las respuestas
        for pregunta in preguntas:
            respuesta = RespuestaEstudiante.objects.create(
                solicitud=solicitud,
                pregunta=pregunta
            )
            
            # Obtener las opciones seleccionadas
            if pregunta.tipo == 'seleccion_multiple':
                opcion_ids = request.POST.getlist(f'pregunta_{pregunta.id}')
                for opcion_id in opcion_ids:
                    opcion = get_object_or_404(OpcionRespuesta, id=opcion_id)
                    respuesta.opciones_seleccionadas.add(opcion)
            elif pregunta.tipo == 'seleccion_unica':
                opcion_id = request.POST.get(f'pregunta_{pregunta.id}')
                if opcion_id:
                    opcion = get_object_or_404(OpcionRespuesta, id=opcion_id)
                    respuesta.opciones_seleccionadas.add(opcion)
            elif pregunta.tipo == 'escritura_libre':
                # Para preguntas de escritura libre, creamos una opción de respuesta con el texto ingresado
                texto_respuesta = request.POST.get(f'pregunta_{pregunta.id}', '')
                if texto_respuesta:
                    # Crear una opción de respuesta para almacenar el texto
                    opcion = OpcionRespuesta.objects.create(
                        pregunta=pregunta,
                        texto=texto_respuesta,
                        orden=0
                    )
                    respuesta.opciones_seleccionadas.add(opcion)
        
        # En lugar de redirigir a la lista de cursos, redirigimos a una página de confirmación
        # que luego redirigirá automáticamente a la lista de cursos
        request.session['solicitud_enviada_curso_id'] = curso_id
        return redirect('principal:solicitud_enviada', curso_id=curso_id)
    
    # Crear formularios dinámicos para cada pregunta
    formularios_preguntas = []
    for pregunta in preguntas:
        form = RespuestaEstudianteForm(pregunta=pregunta)
        formularios_preguntas.append((pregunta, form))
    
    context = {
        'curso': curso,
        'formulario': formulario,
        'formularios_preguntas': formularios_preguntas
    }
    
    return render(request, 'formularios/aplicar_curso.html', context)

# Vista para mostrar la página de confirmación después de enviar una solicitud
@login_required
def solicitud_enviada(request, curso_id):
    """
    Vista para mostrar una página de confirmación después de enviar una solicitud.
    """
    curso = get_object_or_404(Curso, id=curso_id)
    
    # Verificar si el estudiante realmente envió una solicitud para este curso
    solicitud_existente = SolicitudInscripcion.objects.filter(
        curso=curso,
        estudiante=request.user
    ).exists()
    
    if not solicitud_existente:
        # Si no existe una solicitud, redirigir a la lista de cursos
        return redirect('principal:cursos')
    
    # Imprimir mensaje de depuración
    print(f"DEBUG: Mostrando página de confirmación para solicitud del curso {curso.name} (ID: {curso.id})")
    
    return render(request, 'formularios/solicitud_enviada.html', {'curso': curso})

# Vistas para los profesores

class SolicitudesInscripcionListView(LoginRequiredMixin, ProfesorRequiredMixin, ListView):
    """
    Vista para que un profesor vea las solicitudes de inscripción a sus cursos.
    """
    model = SolicitudInscripcion
    template_name = 'formularios/solicitudes_list.html'
    context_object_name = 'solicitudes'
    
    def get_queryset(self):
        # Obtener solo las solicitudes de los cursos que imparte el profesor
        return SolicitudInscripcion.objects.filter(
            curso__teacher=self.request.user,
            estado='pendiente'
        ).order_by('-fecha_solicitud')

class SolicitudInscripcionDetailView(LoginRequiredMixin, ProfesorRequiredMixin, DetailView):
    """
    Vista para que un profesor vea el detalle de una solicitud de inscripción.
    """
    model = SolicitudInscripcion
    template_name = 'formularios/solicitud_detail.html'
    context_object_name = 'solicitud'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        solicitud = self.get_object()
        
        # Verificar que el profesor sea el profesor del curso
        if solicitud.curso.teacher != self.request.user:
            raise PermissionDenied
        
        # Obtener las respuestas del estudiante
        respuestas = solicitud.respuestas.all().select_related('pregunta').prefetch_related('opciones_seleccionadas')
        context['respuestas'] = respuestas
        
        return context

@login_required
def aprobar_solicitud(request, pk):
    """
    Vista para que un profesor apruebe una solicitud de inscripción.
    """
    solicitud = get_object_or_404(SolicitudInscripcion, pk=pk)
    
    # Verificar que el profesor sea el profesor del curso
    if solicitud.curso.teacher != request.user:
        raise PermissionDenied
    
    # Aprobar la solicitud
    matricula = solicitud.aprobar(request.user)
    
    # Enviar correo de confirmación al estudiante
    try:
        nombre_estudiante = solicitud.estudiante.get_full_name() or solicitud.estudiante.username
        nombre_curso = solicitud.curso.name
        email_estudiante = solicitud.estudiante.email
        
        if email_estudiante:
            asunto = f'¡Enhorabuena! Su aplicación al curso {nombre_curso} ha sido aprobada'
            mensaje = f'''¡Enhorabuena! Su aplicación al curso "{nombre_curso}" ha sido aprobada.

Ya puede acceder al curso y comenzar con las actividades académicas.

Saludos cordiales,
Centro Fray Bartolomé de las Casas'''
            
            send_mail(
                asunto,
                mensaje,
                settings.DEFAULT_FROM_EMAIL,
                [email_estudiante],
                fail_silently=False,
            )
            print(f"Correo enviado exitosamente a {email_estudiante} para el curso {nombre_curso}")
        else:
            print(f"No se pudo enviar correo: el estudiante {nombre_estudiante} no tiene email registrado")
    except Exception as e:
        print(f"Error al enviar correo de aprobación: {str(e)}")
        # No interrumpimos el proceso si falla el envío del correo
    
    # Agregar un solo mensaje de éxito
    messages.success(request, f'La solicitud de {solicitud.estudiante.get_full_name() or solicitud.estudiante.username} ha sido aprobada.')
    
    # Verificar si la solicitud viene del perfil o de la página de solicitudes
    referer = request.META.get('HTTP_REFERER', '')
    if 'profile' in referer:
        return redirect('principal:profile')
    else:
        return redirect('principal:solicitudes_list')

@login_required
def rechazar_solicitud(request, pk):
    """
    Vista para que un profesor rechace una solicitud de inscripción.
    """
    solicitud = get_object_or_404(SolicitudInscripcion, pk=pk)
    
    # Verificar que el profesor sea el profesor del curso
    if solicitud.curso.teacher != request.user:
        raise PermissionDenied
    
    # Rechazar la solicitud
    solicitud.rechazar(request.user)
    
    # Enviar correo de notificación al estudiante
    try:
        nombre_estudiante = solicitud.estudiante.get_full_name() or solicitud.estudiante.username
        nombre_curso = solicitud.curso.name
        email_estudiante = solicitud.estudiante.email
        
        if email_estudiante:
            asunto = f'Su aplicación al curso {nombre_curso} ha sido denegada'
            mensaje = f'''Lo sentimos! Su aplicación al curso "{nombre_curso}" ha sido denegada.

Le recomendamos revisar los requisitos del curso y considerar aplicar en futuras convocatorias.

Si tiene alguna pregunta, no dude en contactarnos.

Saludos cordiales,
Centro Fray Bartolomé de las Casas'''
            
            send_mail(
                asunto,
                mensaje,
                settings.DEFAULT_FROM_EMAIL,
                [email_estudiante],
                fail_silently=False,
            )
            print(f"Correo de denegación enviado exitosamente a {email_estudiante} para el curso {nombre_curso}")
        else:
            print(f"No se pudo enviar correo: el estudiante {nombre_estudiante} no tiene email registrado")
    except Exception as e:
        print(f"Error al enviar correo de denegación: {str(e)}")
        # No interrumpimos el proceso si falla el envío del correo
    
    messages.success(request, f'La solicitud de {solicitud.estudiante.get_full_name() or solicitud.estudiante.username} ha sido rechazada.')
    return redirect('principal:solicitudes_list')

@login_required
def guardar_pregunta_y_redirigir(request, formulario_id):
    """
    Vista para guardar una pregunta y redirigir a la página de opciones.
    """
    # Verificar que el usuario pertenezca al grupo 'Secretaria'
    if not request.user.groups.filter(name='Secretaria').exists():
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('principal:formulario_list')
    
    # Obtener el formulario
    formulario = get_object_or_404(FormularioAplicacion, pk=formulario_id)
    
    if request.method == 'POST':
        # Crear una nueva pregunta
        requerida_value = request.POST.get('requerida', 'True')
        requerida = requerida_value.lower() == 'true' if isinstance(requerida_value, str) else bool(requerida_value)
        
        # Imprimir información de depuración
        print(f"Guardando pregunta para formulario {formulario_id}")
        print(f"Texto: {request.POST.get('texto', '')}")
        print(f"Tipo: {request.POST.get('tipo', 'seleccion_multiple')}")
        print(f"Requerida: {requerida}")
        print(f"Orden: {int(request.POST.get('orden', 0))}")
        
        pregunta = PreguntaFormulario(
            formulario=formulario,
            texto=request.POST.get('texto', ''),
            tipo=request.POST.get('tipo', 'seleccion_multiple'),
            requerida=requerida,
            orden=int(request.POST.get('orden', 0))
        )
        pregunta.save()
        
        # Imprimir información de la pregunta guardada
        print(f"Pregunta guardada con ID: {pregunta.pk}")
        
        # Limpiar la caché de la sesión para forzar una recarga de los datos
        if 'cursos_con_formularios' in request.session:
            del request.session['cursos_con_formularios']
        
        # No mostramos mensaje aquí para evitar duplicación, ya que la vista FormularioPreguntasView ya muestra un mensaje
        
        # Usar una redirección con JavaScript
        from django.http import HttpResponse
        redirect_url = reverse('principal:pregunta_opciones', kwargs={'pk': pregunta.pk}) + '?from_redirect=1'
        return HttpResponse(f"""
            <html>
                <head>
                    <title>Redirigiendo...</title>
                    <script>
                        window.location.href = "{redirect_url}";
                    </script>
                </head>
                <body>
                    <p>Redirigiendo a la página de opciones de respuesta...</p>
                    <p>Si no eres redirigido automáticamente, <a href="{redirect_url}">haz clic aquí</a>.</p>
                </body>
            </html>
        """)
    
    # Si no es POST, redirigir a la página de preguntas del formulario
    return redirect('principal:formulario_preguntas', pk=formulario_id)

@login_required
def eliminar_formulario(request, pk):
    """
    Vista para eliminar un formulario de aplicación.
    """
    # Verificar que el usuario pertenezca al grupo Secretaria
    if not request.user.groups.filter(name='Secretaria').exists():
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('principal:cursos')
    
    formulario = get_object_or_404(FormularioAplicacion, pk=pk)
    curso = formulario.curso
    curso_id = curso.id  # Guardar el ID del curso antes de eliminar el formulario
    
    # Imprimir información de depuración antes de eliminar
    print(f"DEBUG: Eliminando formulario {pk} del curso {curso.name} (ID: {curso_id})")
    print(f"DEBUG: Verificando relación antes de eliminar: {FormularioAplicacion.objects.filter(curso=curso).exists()}")
    
    # Eliminar el formulario
    formulario.delete()
    
    # Verificar que se haya eliminado correctamente
    print(f"DEBUG: Verificando relación después de eliminar: {FormularioAplicacion.objects.filter(curso=curso).exists()}")
    
    # Limpiar la caché de la sesión para forzar una recarga de los datos
    if 'cursos_con_formularios' in request.session:
        del request.session['cursos_con_formularios']
    
    messages.success(request, 'El formulario de aplicación ha sido eliminado correctamente.')
    return redirect('principal:cursos')

# Vistas para recuperación de contraseña
import random
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User

def password_reset_request(request):
    """
    Vista para solicitar el restablecimiento de contraseña.
    El usuario ingresa su correo electrónico y se le envía un código de verificación.
    """
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            # Generar código aleatorio de 4 dígitos
            verification_code = str(random.randint(1000, 9999))
            # Almacenar datos temporales en la sesión
            request.session['reset_verification_code'] = verification_code
            request.session['reset_user_id'] = user.id
            
            # Enviar email con el código de verificación
            email_text = f'Para restablecer su contraseña en el Centro Fray Bartolome de las Casas, ingrese el siguiente código: {verification_code}'
            try:
                send_mail(
                    'Código para Restablecer Contraseña - Centro Fray Bartolome de las Casas',
                    email_text,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
                messages.success(request, 'Se ha enviado un código de verificación a su correo electrónico.')
                return redirect('principal:password_reset_verify')
            except Exception as e:
                print(f"Error al enviar email: {str(e)}")
                messages.error(request, 'Error al enviar el código de verificación. Por favor, intente nuevamente más tarde.')
        except User.DoesNotExist:
            messages.error(request, 'No existe una cuenta con ese correo electrónico.')
    
    return render(request, 'registration/password_reset_request.html')

def password_reset_verify(request):
    """
    Vista para verificar el código enviado al correo electrónico.
    """
    if 'reset_verification_code' not in request.session or 'reset_user_id' not in request.session:
        messages.error(request, 'La sesión ha expirado. Por favor, inicie el proceso nuevamente.')
        return redirect('principal:password_reset_request')
    
    if request.method == 'POST':
        code = request.POST.get('code')
        if code == request.session.get('reset_verification_code'):
            return redirect('principal:password_reset_confirm')
        else:
            messages.error(request, 'El código ingresado no es válido. Por favor, intente nuevamente.')
    
    return render(request, 'registration/password_reset_verify.html')

def password_reset_confirm(request):
    """
    Vista para establecer la nueva contraseña después de verificar el código.
    """
    if 'reset_user_id' not in request.session:
        messages.error(request, 'La sesión ha expirado. Por favor, inicie el proceso nuevamente.')
        return redirect('principal:password_reset_request')
    
    if request.method == 'POST':
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        
        if password1 != password2:
            messages.error(request, 'Las contraseñas no coinciden. Por favor, inténtelo nuevamente.')
            return render(request, 'registration/password_reset_confirm.html')
        
        try:
            user = User.objects.get(id=request.session.get('reset_user_id'))
            user.password = make_password(password1)
            user.save()
            
            # Enviar correo de confirmación de cambio de contraseña
            try:
                nombre_usuario = user.get_full_name() or user.username
                email_usuario = user.email
                
                if email_usuario:
                    asunto = 'Su contraseña ha sido cambiada satisfactoriamente'
                    mensaje = f'''Estimado/a {nombre_usuario},

Su contraseña ha sido cambiada satisfactoriamente en el Centro Fray Bartolomé de las Casas.

Si usted no realizó este cambio, por favor contacte inmediatamente con el administrador del sistema.

Fecha y hora del cambio: {timezone.now().strftime('%d/%m/%Y a las %H:%M')}

Saludos cordiales,
Centro Fray Bartolomé de las Casas'''
                    
                    send_mail(
                        asunto,
                        mensaje,
                        settings.DEFAULT_FROM_EMAIL,
                        [email_usuario],
                        fail_silently=False,
                    )
                    print(f"Correo de confirmación de cambio de contraseña enviado a {email_usuario}")
                else:
                    print(f"No se pudo enviar correo: el usuario {nombre_usuario} no tiene email registrado")
            except Exception as e:
                print(f"Error al enviar correo de confirmación de cambio de contraseña: {str(e)}")
                # No interrumpimos el proceso si falla el envío del correo
            
            # Limpiar datos de sesión
            del request.session['reset_verification_code']
            del request.session['reset_user_id']
            
            messages.success(request, 'Su contraseña ha sido restablecida exitosamente. Ahora puede iniciar sesión con su nueva contraseña.')
            return redirect('login')
        except User.DoesNotExist:
            messages.error(request, 'Ha ocurrido un error. Por favor, inicie el proceso nuevamente.')
            return redirect('principal:password_reset_request')
    
    return render(request, 'registration/password_reset_confirm.html')