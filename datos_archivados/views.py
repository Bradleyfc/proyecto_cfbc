from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator
from django.db.models import Count
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.mail import send_mail
from django.conf import settings
import json

def es_secretaria(user):
    """Verifica si el usuario pertenece al grupo Secretaria"""
    return user.groups.filter(name='Secretaria').exists()

def tiene_permisos_datos_archivados(user):
    """Verifica si el usuario tiene permisos para acceder a datos archivados (Secretaria o Admin)"""
    return (user.groups.filter(name='Secretaria').exists() or 
            user.is_superuser or 
            user.is_staff)

def permisos_datos_archivados_required(view_func):
    """Decorador personalizado que verifica permisos para datos archivados"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        
        if not tiene_permisos_datos_archivados(request.user):
            return render(request, 'datos_archivados/sin_permisos.html', status=403)
        
        return view_func(request, *args, **kwargs)
    return wrapper

class PermisosDataArchivadosRequiredMixin:
    """Mixin que verifica permisos para datos archivados (Secretaria o Admin)"""
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        
        if not tiene_permisos_datos_archivados(request.user):
            return render(request, 'datos_archivados/sin_permisos.html', status=403)
        
        return super().dispatch(request, *args, **kwargs)

@method_decorator(login_required, name='dispatch')
class DashboardArchivadosView(TemplateView):
    """Vista principal del dashboard de datos archivados"""
    template_name = 'datos_archivados/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Verificar permisos
        if not tiene_permisos_datos_archivados(self.request.user):
            context['sin_permisos'] = True
            return context
        
        # Importar modelos aquí para evitar problemas de importación
        try:
            from .models import DatoArchivadoDinamico, MigracionLog
            
            # Estadísticas actuales basadas en datos existentes
            total_datos_actuales = DatoArchivadoDinamico.objects.count()
            total_tablas_actuales = DatoArchivadoDinamico.objects.values('tabla_origen').distinct().count()
            ultima_migracion = MigracionLog.objects.first()
            
            # Calcular estadísticas dinámicas
            if total_datos_actuales > 0:
                # Hay datos: usar estadísticas reales
                tablas_inspeccionadas = ultima_migracion.tablas_inspeccionadas if ultima_migracion else 0
                tablas_con_datos = total_tablas_actuales  # Basado en datos actuales
                tablas_vacias = max(0, tablas_inspeccionadas - tablas_con_datos)  # Calculado dinámicamente
            else:
                # No hay datos: todo en cero
                tablas_inspeccionadas = 0
                tablas_con_datos = 0
                tablas_vacias = 0
            
            # Estadísticas básicas
            context.update({
                'total_datos_archivados': total_datos_actuales,
                'total_tablas_migradas': tablas_con_datos,
                'ultima_migracion': ultima_migracion,
                'tablas_inspeccionadas_actuales': tablas_inspeccionadas,
                'tablas_vacias_actuales': tablas_vacias,
            })
            
            # Distribución por tabla (primeras 10)
            context['datos_por_tabla'] = DatoArchivadoDinamico.objects.values(
                'tabla_origen'
            ).annotate(
                total=Count('id')
            ).order_by('-total')[:10]
            
            # Últimas migraciones (primeras 5)
            context['ultimas_migraciones'] = MigracionLog.objects.all()[:5]
            
        except Exception as e:
            # Si hay problemas con los modelos, usar valores por defecto
            context.update({
                'total_datos_archivados': 0,
                'total_tablas_migradas': 0,
                'ultima_migracion': None,
                'datos_por_tabla': [],
                'ultimas_migraciones': [],
            })
        
        return context

@login_required
def configurar_migracion_view(request):
    """Vista para configurar y ejecutar migración automática desde MariaDB"""
    # Verificar permisos
    if not tiene_permisos_datos_archivados(request.user):
        return render(request, 'datos_archivados/sin_permisos.html', status=403)
    
    if request.method == 'POST':
        # Obtener datos de configuración
        host = request.POST.get('host')
        database = request.POST.get('database')
        user = request.POST.get('user')
        password = request.POST.get('password')
        port = request.POST.get('port', 3306)
        
        if not all([host, database, user, password]):
            messages.error(request, 'Todos los campos de conexión son obligatorios.')
            return render(request, 'datos_archivados/configurar_migracion.html')
        
        try:
            # Importar servicio aquí para evitar problemas de importación
            from .services import MigracionService
            
            # Crear servicio de migración
            servicio = MigracionService(host, database, user, password, int(port))
            
            # Probar conexión
            if not servicio.conectar_mariadb():
                messages.error(request, 'No se pudo conectar a la base de datos MariaDB. Verifique los datos de conexión.')
                return render(request, 'datos_archivados/configurar_migracion.html')
            
            servicio.desconectar_mariadb()
            
            # Ejecutar migración automática en hilo separado
            def ejecutar_migracion_automatica():
                try:
                    servicio.inspeccionar_y_migrar_automaticamente(request.user)
                except Exception as e:
                    pass  # Error handling is in the service
            
            import threading
            thread = threading.Thread(target=ejecutar_migracion_automatica)
            thread.daemon = True
            thread.start()
            
            messages.success(request, 'Migración automática iniciada correctamente. El sistema inspeccionará la base de datos y migrará todos los datos automáticamente.')
            return redirect('datos_archivados:dashboard')
            
        except Exception as e:
            messages.error(request, f'Error al iniciar la migración: {str(e)}')
    
    return render(request, 'datos_archivados/configurar_migracion.html')

@login_required
def detalle_dato_archivado(request, pk):
    """Vista para mostrar el detalle completo de un dato archivado"""
    if not tiene_permisos_datos_archivados(request.user):
        return render(request, 'datos_archivados/sin_permisos.html', status=403)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.shortcuts import get_object_or_404
        import json
        
        # Obtener el dato archivado
        dato = get_object_or_404(DatoArchivadoDinamico, pk=pk)
        
        # Preparar los datos para mostrar
        context = {
            'dato': dato,
            'datos_formateados': json.dumps(dato.datos_originales, indent=2, ensure_ascii=False),
            'estructura_formateada': json.dumps(
                json.loads(dato.estructura_tabla) if dato.estructura_tabla else {}, 
                indent=2, ensure_ascii=False
            ) if dato.estructura_tabla else None,
        }
        
        return render(request, 'datos_archivados/dato_detail.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar el detalle: {str(e)}')
        return redirect('datos_archivados:datos_list')

@method_decorator(login_required, name='dispatch')
class TablasArchivadosListView(TemplateView):
    """Vista para listar las tablas archivadas disponibles"""
    template_name = 'datos_archivados/tablas_list.html'
    
    def dispatch(self, request, *args, **kwargs):
        if not tiene_permisos_datos_archivados(request.user):
            return render(request, 'datos_archivados/sin_permisos.html', status=403)
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        try:
            from .models import DatoArchivadoDinamico
            from django.db.models import Count, Max, Q
            
            # Obtener parámetros de búsqueda
            search_query = self.request.GET.get('search', '').strip()
            search_type = self.request.GET.get('search_type', 'tabla')
            
            # Query base
            queryset = DatoArchivadoDinamico.objects.all()
            
            # Aplicar filtros de búsqueda
            if search_query:
                if search_type == 'tabla':
                    # Buscar por nombre de tabla
                    queryset = queryset.filter(tabla_origen__icontains=search_query)
                elif search_type == 'contenido':
                    # Buscar en el contenido de los datos (JSON)
                    queryset = queryset.filter(
                        Q(datos_originales__icontains=search_query) |
                        Q(nombre_registro__icontains=search_query)
                    )
                elif search_type == 'global':
                    # Búsqueda global (tabla + contenido)
                    queryset = queryset.filter(
                        Q(tabla_origen__icontains=search_query) |
                        Q(datos_originales__icontains=search_query) |
                        Q(nombre_registro__icontains=search_query)
                    )
            
            # Obtener estadísticas por tabla con filtros aplicados
            tablas_stats = queryset.values('tabla_origen').annotate(
                total_registros=Count('id'),
                ultima_migracion=Max('fecha_migracion')
            ).order_by('tabla_origen')
            
            # Estadísticas generales
            total_registros_filtrados = queryset.count()
            total_registros_totales = DatoArchivadoDinamico.objects.count()
            
            context.update({
                'tablas_stats': tablas_stats,
                'total_tablas': tablas_stats.count(),
                'total_registros': total_registros_totales,
                'total_registros_filtrados': total_registros_filtrados,
                'search_query': search_query,
                'search_type': search_type,
                'is_filtered': bool(search_query),
            })
            
        except Exception as e:
            context.update({
                'tablas_stats': [],
                'total_tablas': 0,
                'total_registros': 0,
                'total_registros_filtrados': 0,
                'search_query': '',
                'search_type': 'tabla',
                'is_filtered': False,
            })
        
        return context

@method_decorator(login_required, name='dispatch')
class DatosArchivadosListView(TemplateView):
    """Vista para listar datos archivados de una tabla específica"""
    template_name = 'datos_archivados/datos_list.html'
    
    def dispatch(self, request, *args, **kwargs):
        if not tiene_permisos_datos_archivados(request.user):
            return render(request, 'datos_archivados/sin_permisos.html', status=403)
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        try:
            from .models import DatoArchivadoDinamico
            
            # Obtener la tabla desde la URL
            tabla = kwargs.get('tabla')
            
            if not tabla:
                # Si no hay tabla, redirigir a la lista de tablas
                context['datos'] = []
                context['tabla_actual'] = None
                return context
            
            # Obtener datos de la tabla específica
            queryset = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla)
            
            # Filtro de búsqueda
            search = self.request.GET.get('search')
            if search:
                queryset = queryset.filter(nombre_registro__icontains=search)
            
            # Ordenamiento
            order_by = self.request.GET.get('order_by', 'fecha_migracion')
            order_direction = self.request.GET.get('order_direction', 'desc')
            
            # Validar campos de ordenamiento
            valid_order_fields = {
                'fecha_migracion': 'fecha_migracion',
                'id_original': 'id_original',
                'nombre': 'nombre_registro',
                'tabla': 'tabla_origen'
            }
            
            if order_by in valid_order_fields:
                order_field = valid_order_fields[order_by]
                if order_direction == 'desc':
                    order_field = f'-{order_field}'
                queryset = queryset.order_by(order_field)
            else:
                # Ordenamiento por defecto
                queryset = queryset.order_by('-fecha_migracion')
            
            # Para ordenamiento por nombre, necesitamos ordenar después de obtener los datos
            # porque nombre_registro puede ser None y necesitamos usar obtener_nombre_legible()
            if order_by == 'nombre':
                datos_list = list(queryset)
                datos_list.sort(
                    key=lambda x: (x.obtener_nombre_legible() or '').lower(),
                    reverse=(order_direction == 'desc')
                )
                context['datos'] = datos_list
            else:
                context['datos'] = queryset
            
            context.update({
                'tabla_actual': tabla,
                'total_registros': queryset.count(),
                'search_query': search or '',
                'order_by': order_by,
                'order_direction': order_direction,
            })
            
        except Exception as e:
            context['datos'] = []
            context['tabla_actual'] = None
            context['total_registros'] = 0
        
        return context

@login_required
def estado_migracion_ajax(request):
    """Vista AJAX para obtener el estado de la migración actual"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        from .models import MigracionLog
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        # Función helper para manejar valores None
        def safe_int(value):
            return value if value is not None else 0
        
        def safe_str(value):
            return value if value is not None else ''
        
        # Función helper para obtener atributos que pueden no existir
        def safe_getattr(obj, attr, default=0):
            return getattr(obj, attr, default) if hasattr(obj, attr) else default
        
        # Buscar migración en progreso
        migracion_en_progreso = MigracionLog.objects.filter(
            estado__in=['iniciada', 'en_progreso']
        ).first()
        
        if migracion_en_progreso:
            # Calcular total de registros migrados
            total_migrados = (
                safe_int(migracion_en_progreso.usuarios_migrados) +
                safe_int(migracion_en_progreso.cursos_academicos_migrados) +
                safe_int(migracion_en_progreso.cursos_migrados) +
                safe_int(migracion_en_progreso.matriculas_migradas) +
                safe_int(migracion_en_progreso.calificaciones_migradas) +
                safe_int(migracion_en_progreso.notas_migradas) +
                safe_int(migracion_en_progreso.asistencias_migradas)
            )
            
            data = {
                'en_progreso': True,
                'estado': safe_str(migracion_en_progreso.estado) or 'iniciada',
                'fecha_inicio': migracion_en_progreso.fecha_inicio.isoformat() if migracion_en_progreso.fecha_inicio else None,
                'total_migrados': total_migrados,
                'tablas_inspeccionadas': safe_int(safe_getattr(migracion_en_progreso, 'tablas_inspeccionadas', 0)),
                'tablas_con_datos': safe_int(safe_getattr(migracion_en_progreso, 'tablas_con_datos', 0)),
                'tablas_vacias': safe_int(safe_getattr(migracion_en_progreso, 'tablas_vacias', 0)),
                'host_origen': safe_str(safe_getattr(migracion_en_progreso, 'host_origen', '')) or 'N/A',
                'base_datos_origen': safe_str(safe_getattr(migracion_en_progreso, 'base_datos_origen', '')) or 'N/A',
            }
        else:
            # Verificar si hay una migración completada recientemente (últimos 10 minutos)
            hace_10_minutos = timezone.now() - timedelta(minutes=10)
            migracion_reciente = MigracionLog.objects.filter(
                fecha_inicio__gte=hace_10_minutos,
                estado='completada'
            ).first()
            
            if migracion_reciente:
                # Calcular total de registros migrados
                total_migrados = (
                    safe_int(migracion_reciente.usuarios_migrados) +
                    safe_int(migracion_reciente.cursos_academicos_migrados) +
                    safe_int(migracion_reciente.cursos_migrados) +
                    safe_int(migracion_reciente.matriculas_migradas) +
                    safe_int(migracion_reciente.calificaciones_migradas) +
                    safe_int(migracion_reciente.notas_migradas) +
                    safe_int(migracion_reciente.asistencias_migradas)
                )
                
                data = {
                    'en_progreso': False,
                    'completada_recientemente': True,
                    'estado': safe_str(migracion_reciente.estado) or 'completada',
                    'fecha_inicio': migracion_reciente.fecha_inicio.isoformat() if migracion_reciente.fecha_inicio else None,
                    'fecha_fin': migracion_reciente.fecha_fin.isoformat() if migracion_reciente.fecha_fin else None,
                    'total_migrados': total_migrados,
                    'tablas_inspeccionadas': safe_int(safe_getattr(migracion_reciente, 'tablas_inspeccionadas', 0)),
                    'tablas_con_datos': safe_int(safe_getattr(migracion_reciente, 'tablas_con_datos', 0)),
                    'tablas_vacias': safe_int(safe_getattr(migracion_reciente, 'tablas_vacias', 0)),
                    'host_origen': safe_str(safe_getattr(migracion_reciente, 'host_origen', '')) or 'N/A',
                }
            else:
                data = {
                    'en_progreso': False,
                    'completada_recientemente': False,
                    'estado': 'sin_migraciones',
                    'mensaje': 'No hay migraciones registradas'
                }
        
        return JsonResponse(data)
    except Exception as e:
        # Log del error para debugging
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error en estado_migracion_ajax: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        
        # Respuesta de error segura
        return JsonResponse({
            'error': 'Error interno del servidor',
            'en_progreso': False,
            'completada_recientemente': False,
            'estado': 'error',
            'mensaje': 'Error al obtener estado de migración'
        })

@login_required
def exportar_excel(request, pk):
    """Vista para exportar un dato archivado a Excel"""
    if not tiene_permisos_datos_archivados(request.user):
        return render(request, 'datos_archivados/sin_permisos.html', status=403)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.shortcuts import get_object_or_404
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import json
        from datetime import datetime
        
        # Obtener el dato archivado
        dato = get_object_or_404(DatoArchivadoDinamico, pk=pk)
        
        # Crear workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Dato_{dato.tabla_origen}_{dato.id_original}"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        info_font = Font(bold=True, color="000000")
        info_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        
        # Título principal
        ws.merge_cells('A1:C1')
        ws['A1'] = f"DETALLE DEL DATO ARCHIVADO"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Información general
        row = 3
        ws[f'A{row}'] = "INFORMACIÓN GENERAL"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        info_data = [
            ("Tabla de Origen", dato.tabla_origen),
            ("ID Original", dato.id_original),
            ("Fecha de Migración", dato.fecha_migracion.strftime('%d/%m/%Y %H:%M:%S')),
            ("Nombre del Registro", dato.obtener_nombre_legible()),
        ]
        
        for campo, valor in info_data:
            ws[f'A{row}'] = campo
            ws[f'A{row}'].font = info_font
            ws[f'A{row}'].fill = info_fill
            ws[f'B{row}'] = str(valor)
            row += 1
        
        # Datos del registro
        row += 1
        ws[f'A{row}'] = "DATOS DEL REGISTRO"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Campo"
        ws[f'B{row}'] = "Valor"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        
        # Agregar datos del registro
        for campo, valor in dato.datos_originales.items():
            row += 1
            ws[f'A{row}'] = str(campo)
            ws[f'B{row}'] = str(valor) if valor is not None else "NULL"
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"dato_archivado_{dato.tabla_origen}_{dato.id_original}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar workbook en response
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar a Excel: {str(e)}')
        return redirect('datos_archivados:dato_detail', pk=pk)

@login_required
def exportar_tabla_excel(request, tabla):
    """Vista para exportar todos los datos de una tabla a Excel"""
    if not tiene_permisos_datos_archivados(request.user):
        return render(request, 'datos_archivados/sin_permisos.html', status=403)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        # Obtener todos los datos de la tabla
        datos = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla).order_by('id_original')
        
        if not datos.exists():
            messages.error(request, f'No hay datos para exportar de la tabla {tabla}.')
            return redirect('datos_archivados:datos_list', tabla=tabla)
        
        # Crear workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Tabla_{tabla}"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=16, color="000000")
        info_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título principal
        ws.merge_cells('A1:F1')
        ws['A1'] = f"DATOS ARCHIVADOS - TABLA: {tabla.upper()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Información de la exportación
        row = 3
        ws[f'A{row}'] = f"Fecha de exportación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws[f'A{row}'].font = info_font
        row += 1
        ws[f'A{row}'] = f"Total de registros: {datos.count()}"
        ws[f'A{row}'].font = info_font
        row += 1
        ws[f'A{row}'] = f"Usuario: {request.user.username}"
        ws[f'A{row}'].font = info_font
        
        # Obtener todos los campos únicos de todos los registros
        todos_los_campos = set()
        for dato in datos:
            todos_los_campos.update(dato.datos_originales.keys())
        
        campos_ordenados = sorted(list(todos_los_campos))
        
        # Headers de la tabla
        row = 6
        ws[f'A{row}'] = "ID Sistema"
        ws[f'B{row}'] = "ID Original"
        ws[f'C{row}'] = "Fecha Migración"
        
        # Agregar headers de campos dinámicos
        col_start = 4  # Columna D
        for i, campo in enumerate(campos_ordenados):
            col_letter = get_column_letter(col_start + i)
            ws[f'{col_letter}{row}'] = campo
        
        # Aplicar estilo a headers
        for col in range(1, len(campos_ordenados) + 4):
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}{row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Agregar datos
        for dato in datos:
            row += 1
            ws[f'A{row}'] = dato.pk
            ws[f'B{row}'] = dato.id_original
            ws[f'C{row}'] = dato.fecha_migracion.strftime('%d/%m/%Y %H:%M')
            
            # Agregar datos de campos dinámicos
            for i, campo in enumerate(campos_ordenados):
                col_letter = get_column_letter(col_start + i)
                valor = dato.datos_originales.get(campo)
                ws[f'{col_letter}{row}'] = str(valor) if valor is not None else "NULL"
            
            # Aplicar bordes a toda la fila
            for col in range(1, len(campos_ordenados) + 4):
                col_letter = get_column_letter(col)
                ws[f'{col_letter}{row}'].border = border
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)  # Máximo 30 caracteres
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)  # Mínimo 10
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"tabla_{tabla}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar workbook en response
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al exportar tabla a Excel: {str(e)}')
        return redirect('datos_archivados:datos_list', tabla=tabla)

@login_required
def eliminar_tablas(request):
    """Vista para eliminar tablas seleccionadas y todos sus registros"""
    if not tiene_permisos_datos_archivados(request.user):
        return render(request, 'datos_archivados/sin_permisos.html', status=403)
    
    if request.method == 'POST':
        try:
            from .models import DatoArchivadoDinamico
            
            # Obtener tablas a eliminar
            tablas = request.POST.getlist('tablas')
            
            if not tablas:
                messages.warning(request, 'No se seleccionaron tablas para eliminar.')
                return redirect('datos_archivados:tablas_list')
            
            # Contar registros antes de eliminar
            total_registros = 0
            for tabla in tablas:
                count = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla).count()
                total_registros += count
            
            # Eliminar registros de las tablas seleccionadas
            eliminados = 0
            for tabla in tablas:
                eliminados += DatoArchivadoDinamico.objects.filter(tabla_origen=tabla).delete()[0]
            
            messages.success(
                request, 
                f'Se eliminaron {len(tablas)} tabla(s) con un total de {eliminados} registros.'
            )
            
        except Exception as e:
            messages.error(request, f'Error al eliminar tablas: {str(e)}')
    
    return redirect('datos_archivados:tablas_list')

def reclamar_usuario_archivado(request):
    """Vista para que usuarios reclamen su cuenta archivada - Paso 1: Validar datos y enviar código"""
    from .forms import ReclamarUsuarioArchivadoForm
    from .models import CodigoVerificacionReclamacion, UsuarioArchivado
    from django.core.mail import send_mail
    from django.conf import settings
    from django.utils import timezone
    from datetime import timedelta
    import random
    import string
    
    if request.method == 'POST':
        form = ReclamarUsuarioArchivadoForm(request.POST)
        if form.is_valid():
            try:
                # Obtener el usuario encontrado y su tipo
                usuario_encontrado = form.cleaned_data['usuario_encontrado']
                tipo_fuente = form.cleaned_data['tipo_fuente']
                
                # Extraer email y username según el tipo de fuente
                if tipo_fuente == 'usuario_archivado':
                    email = usuario_encontrado.email
                    username = usuario_encontrado.username
                    usuario_archivado_id = usuario_encontrado.id
                    dato_archivado_id = None
                else:  # dato_dinamico
                    datos = usuario_encontrado.datos_originales
                    email = datos.get('email')
                    username = datos.get('username')
                    usuario_archivado_id = None
                    dato_archivado_id = usuario_encontrado.id
                
                # Generar código de 4 dígitos
                codigo = ''.join(random.choices(string.digits, k=4))
                
                # Crear o actualizar código de verificación
                fecha_expiracion = timezone.now() + timedelta(minutes=15)
                
                # Eliminar códigos anteriores para este email
                CodigoVerificacionReclamacion.objects.filter(email=email).delete()
                
                # Crear nuevo código
                codigo_obj = CodigoVerificacionReclamacion.objects.create(
                    email=email,
                    codigo=codigo,
                    fecha_expiracion=fecha_expiracion
                )
                
                # Vincular con la fuente correspondiente
                if tipo_fuente == 'usuario_archivado':
                    codigo_obj.usuario_archivado = usuario_encontrado
                else:
                    codigo_obj.dato_archivado = usuario_encontrado
                codigo_obj.save()
                
                # Enviar email
                email_text = f'''Hola {username},

Se ha solicitado reclamar su cuenta archivada en el Centro Fray Bartolomé de las Casas.

Para continuar con el proceso, ingrese el siguiente código de verificación:

{codigo}

Este código expirará en 15 minutos.

Si usted no solicitó esta acción, ignore este mensaje.

Centro Fray Bartolomé de las Casas'''
                
                send_mail(
                    'Código de Verificación - Reclamación de Cuenta Archivada',
                    email_text,
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
                
                # Guardar información en sesión para el siguiente paso
                request.session['reclamacion_tipo_fuente'] = tipo_fuente
                if usuario_archivado_id:
                    request.session['reclamacion_usuario_archivado_id'] = usuario_archivado_id
                if dato_archivado_id:
                    request.session['reclamacion_dato_archivado_id'] = dato_archivado_id
                request.session['reclamacion_email'] = email
                request.session['reclamacion_nueva_password'] = form.cleaned_data['nueva_password']
                
                messages.success(request, f'Se ha enviado un código de verificación a {email}')
                return redirect('datos_archivados:verificar_codigo_reclamacion_tradicional')
                
            except Exception as e:
                messages.error(request, f'Error al enviar el código de verificación: {str(e)}')
    else:
        form = ReclamarUsuarioArchivadoForm()
    
    return render(request, 'datos_archivados/reclamar_usuario.html', {'form': form})

def reenviar_codigo_reclamacion_tradicional(request):
    """Vista para reenviar el código de verificación"""
    from .models import CodigoVerificacionReclamacion, UsuarioArchivado, DatoArchivadoDinamico
    from django.core.mail import send_mail
    from django.conf import settings
    from django.utils import timezone
    from datetime import timedelta
    import random
    import string
    
    # Verificar que hay una reclamación en proceso
    tipo_fuente = request.session.get('reclamacion_tipo_fuente')
    usuario_archivado_id = request.session.get('reclamacion_usuario_archivado_id')
    dato_archivado_id = request.session.get('reclamacion_dato_archivado_id')
    email = request.session.get('reclamacion_email')
    
    if not email or not tipo_fuente:
        messages.error(request, 'No hay un proceso de reclamación activo.')
        return redirect('datos_archivados:reclamar_usuario')
    
    try:
        # Generar nuevo código de 4 dígitos
        codigo = ''.join(random.choices(string.digits, k=4))
        
        # Eliminar códigos anteriores para este email
        CodigoVerificacionReclamacion.objects.filter(email=email).delete()
        
        # Crear nuevo código
        fecha_expiracion = timezone.now() + timedelta(minutes=15)
        codigo_obj = CodigoVerificacionReclamacion.objects.create(
            email=email,
            codigo=codigo,
            fecha_expiracion=fecha_expiracion
        )
        
        # Vincular con la fuente correspondiente
        if tipo_fuente == 'usuario_archivado' and usuario_archivado_id:
            usuario_archivado = UsuarioArchivado.objects.get(id=usuario_archivado_id)
            codigo_obj.usuario_archivado = usuario_archivado
            username = usuario_archivado.username
        elif tipo_fuente == 'dato_dinamico' and dato_archivado_id:
            dato_archivado = DatoArchivadoDinamico.objects.get(id=dato_archivado_id)
            codigo_obj.dato_archivado = dato_archivado
            datos = dato_archivado.datos_originales
            username = datos.get('username', email.split('@')[0])
        else:
            username = email.split('@')[0]
        
        codigo_obj.save()
        
        # Enviar email
        email_text = f'''Hola {username},

Se ha solicitado un nuevo código de verificación para reclamar su cuenta archivada en el Centro Fray Bartolomé de las Casas.

Su nuevo código de verificación es:

{codigo}

Este código expirará en 15 minutos.

Si usted no solicitó esta acción, ignore este mensaje.

Centro Fray Bartolomé de las Casas'''
        
        send_mail(
            'Nuevo Código de Verificación - Reclamación de Cuenta',
            email_text,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )
        
        messages.success(request, f'Se ha enviado un nuevo código de verificación a {email}')
        return redirect('datos_archivados:verificar_codigo_reclamacion_tradicional')
        
    except Exception as e:
        messages.error(request, f'Error al reenviar el código: {str(e)}')
        return redirect('datos_archivados:verificar_codigo_reclamacion_tradicional')

def verificar_codigo_reclamacion_tradicional(request):
    """Vista para verificar el código de 4 dígitos en reclamación tradicional"""
    from .models import CodigoVerificacionReclamacion, UsuarioArchivado, DatoArchivadoDinamico
    from django.contrib.auth import login
    from django.contrib.auth.models import User, Group
    
    # Verificar que hay una reclamación en proceso
    tipo_fuente = request.session.get('reclamacion_tipo_fuente')
    usuario_archivado_id = request.session.get('reclamacion_usuario_archivado_id')
    dato_archivado_id = request.session.get('reclamacion_dato_archivado_id')
    email = request.session.get('reclamacion_email')
    nueva_password = request.session.get('reclamacion_nueva_password')
    
    if not email or not nueva_password or not tipo_fuente:
        messages.error(request, 'No hay un proceso de reclamación activo.')
        return redirect('datos_archivados:reclamar_usuario')
    
    if request.method == 'POST':
        codigo_ingresado = request.POST.get('code', '').strip()
        
        if not codigo_ingresado:
            messages.error(request, 'Debe ingresar el código de verificación.')
            return render(request, 'datos_archivados/verificar_codigo_reclamacion_tradicional.html', {
                'email': email
            })
        
        try:
            # Buscar código válido (el más reciente si hay múltiples)
            codigo_verificacion = CodigoVerificacionReclamacion.objects.filter(
                email=email,
                codigo=codigo_ingresado,
                usado=False
            ).order_by('-fecha_creacion').first()
            
            if not codigo_verificacion:
                raise CodigoVerificacionReclamacion.DoesNotExist()
            
            if codigo_verificacion.is_expired():
                messages.error(request, 'El código de verificación ha expirado. Solicite uno nuevo.')
                return render(request, 'datos_archivados/verificar_codigo_reclamacion_tradicional.html', {
                    'email': email
                })
            
            # NO marcar como usado todavía - esperar a que todo se complete
            
            # Obtener datos según el tipo de fuente
            if tipo_fuente == 'usuario_archivado':
                usuario_archivado = UsuarioArchivado.objects.get(id=usuario_archivado_id)
                username_original = usuario_archivado.username
                email_usuario = usuario_archivado.email
                first_name = usuario_archivado.first_name
                last_name = usuario_archivado.last_name
            else:  # dato_dinamico
                dato_archivado = DatoArchivadoDinamico.objects.get(id=dato_archivado_id)
                datos = dato_archivado.datos_originales
                username_original = datos.get('username', '')
                email_usuario = datos.get('email', '')
                first_name = datos.get('first_name', '')
                last_name = datos.get('last_name', '')
            
            # Generar username único con protección contra condiciones de carrera
            from django.db import IntegrityError
            
            user = None
            max_intentos = 100
            
            for intento in range(max_intentos):
                # Generar username candidato
                if intento == 0:
                    username_candidato = username_original
                else:
                    username_candidato = f"{username_original}{intento - 1}"
                
                # Verificar si ya existe
                if User.objects.filter(username=username_candidato).exists():
                    continue
                
                # Intentar crear el usuario
                try:
                    user = User.objects.create_user(
                        username=username_candidato,
                        email=email_usuario,
                        first_name=first_name,
                        last_name=last_name,
                        password=nueva_password,
                        is_active=True
                    )
                    username_final = username_candidato
                    break  # Usuario creado exitosamente
                except IntegrityError:
                    # El username fue creado por otro proceso entre la verificación y la creación
                    # Continuar con el siguiente número
                    continue
            
            if not user:
                messages.error(request, 'No se pudo generar un username único. Contacte al administrador.')
                return redirect('datos_archivados:reclamar_usuario')
            
            # Asignar al grupo Estudiantes
            try:
                grupo_estudiantes = Group.objects.get(name='Estudiantes')
                user.groups.add(grupo_estudiantes)
            except Group.DoesNotExist:
                grupo_estudiantes = Group.objects.create(name='Estudiantes')
                user.groups.add(grupo_estudiantes)
            
            # Vincular con la fuente correspondiente
            if tipo_fuente == 'usuario_archivado':
                usuario_archivado.usuario_actual = user
                usuario_archivado.save()
            
            # Enviar email de confirmación
            username_cambio = username_original != username_final
            
            if username_cambio:
                mensaje_username = f'''IMPORTANTE: Su nombre de usuario ha cambiado
Nombre de usuario original: {username_original}
Nuevo nombre de usuario: {username_final}

Esto se debe a que ya existía un usuario con el nombre "{username_original}" en el sistema.
Por favor, use "{username_final}" para futuros inicios de sesión.'''
            else:
                mensaje_username = f'Su nombre de usuario es: {username_final}'
            
            mensaje = f'''¡Bienvenido de vuelta al Centro Fray Bartolomé de las Casas!

Su cuenta ha sido reactivada exitosamente.

DETALLES DE SU CUENTA:
{mensaje_username}
Correo electrónico: {user.email}
Nombre completo: {user.get_full_name()}

Ahora puede acceder a todos los servicios del sistema usando sus credenciales.

Si tiene alguna pregunta o necesita ayuda, no dude en contactarnos.

Saludos cordiales,
Centro Fray Bartolomé de las Casas'''
            
            send_mail(
                'Cuenta Reactivada - Centro Fray Bartolomé de las Casas',
                mensaje,
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
            )
            
            # Marcar código como usado SOLO después de que todo se completó exitosamente
            codigo_verificacion.usado = True
            codigo_verificacion.save()
            
            # Limpiar sesión
            if 'reclamacion_usuario_archivado_id' in request.session:
                del request.session['reclamacion_usuario_archivado_id']
            if 'reclamacion_dato_archivado_id' in request.session:
                del request.session['reclamacion_dato_archivado_id']
            if 'reclamacion_tipo_fuente' in request.session:
                del request.session['reclamacion_tipo_fuente']
            del request.session['reclamacion_email']
            del request.session['reclamacion_nueva_password']
            
            # Iniciar sesión automáticamente con backend específico
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            
            messages.success(
                request, 
                f'¡Bienvenido de vuelta, {user.get_full_name() or user.username}! '
                'Su cuenta ha sido reactivada exitosamente. '
                'Se le ha enviado un email con los detalles de su cuenta.'
            )
            
            return redirect('principal:login_redirect')
            
        except CodigoVerificacionReclamacion.DoesNotExist:
            messages.error(request, 'Código de verificación incorrecto.')
            return render(request, 'datos_archivados/verificar_codigo_reclamacion_tradicional.html', {
                'email': email
            })
        except Exception as e:
            messages.error(request, f'Error al reactivar la cuenta: {str(e)}')
            return redirect('datos_archivados:reclamar_usuario')
    
    return render(request, 'datos_archivados/verificar_codigo_reclamacion_tradicional.html', {
        'email': email
    })

def buscar_usuario_archivado(request):
    """Vista para buscar usuarios archivados disponibles para reclamar"""
    from .forms import BuscarUsuarioArchivadoForm
    
    form = BuscarUsuarioArchivadoForm()
    usuarios_encontrados = []
    
    if request.method == 'GET' and 'busqueda' in request.GET:
        form = BuscarUsuarioArchivadoForm(request.GET)
        if form.is_valid():
            usuarios_encontrados = form.buscar_usuarios()
    
    context = {
        'form': form,
        'usuarios_encontrados': usuarios_encontrados,
    }
    
    return render(request, 'datos_archivados/buscar_usuario.html', context)

def iniciar_reclamacion_usuario(request, dato_id):
    """Vista para iniciar el proceso de reclamación enviando código de verificación"""
    from .models import DatoArchivadoDinamico, CodigoVerificacionReclamacion
    from django.shortcuts import get_object_or_404
    from django.core.mail import send_mail
    from django.conf import settings
    from django.utils import timezone
    from datetime import timedelta
    import random
    import string
    
    # Obtener el dato archivado
    dato = get_object_or_404(DatoArchivadoDinamico, id=dato_id)
    datos = dato.datos_originales
    email = datos.get('email') or datos.get('correo')
    
    if not email:
        messages.error(request, 'No se encontró un email válido para este usuario archivado.')
        return redirect('datos_archivados:buscar_usuario')
    
    # Generar código de 4 dígitos
    codigo = ''.join(random.choices(string.digits, k=4))
    
    # Crear o actualizar código de verificación
    fecha_expiracion = timezone.now() + timedelta(minutes=15)  # Expira en 15 minutos
    
    # Eliminar códigos anteriores para este email
    CodigoVerificacionReclamacion.objects.filter(email=email).delete()
    
    # Crear nuevo código
    codigo_verificacion = CodigoVerificacionReclamacion.objects.create(
        email=email,
        codigo=codigo,
        dato_archivado=dato,
        fecha_expiracion=fecha_expiracion
    )
    
    # Enviar email
    username = datos.get('username') or datos.get('user') or 'Usuario'
    email_text = f'''Hola {username},

Se ha solicitado reclamar su cuenta archivada en el Centro Fray Bartolomé de las Casas.

Para continuar con el proceso, ingrese el siguiente código de verificación:

{codigo}

Este código expirará en 15 minutos.

Si usted no solicitó esta acción, ignore este mensaje.

Centro Fray Bartolomé de las Casas'''
    
    try:
        send_mail(
            'Código de Verificación - Reclamación de Cuenta Archivada',
            email_text,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )
        
        # Guardar información en sesión para el siguiente paso
        request.session['reclamacion_dato_id'] = dato_id
        request.session['reclamacion_email'] = email
        
        messages.success(request, f'Se ha enviado un código de verificación a {email}')
        return redirect('datos_archivados:verificar_codigo_reclamacion')
        
    except Exception as e:
        messages.error(request, f'Error al enviar el código de verificación: {str(e)}')
        return redirect('datos_archivados:buscar_usuario')

def reenviar_codigo_reclamacion(request):
    """Vista para reenviar el código de verificación (búsqueda por email)"""
    from .models import CodigoVerificacionReclamacion, DatoArchivadoDinamico
    from django.core.mail import send_mail
    from django.conf import settings
    from django.utils import timezone
    from datetime import timedelta
    import random
    import string
    
    # Verificar que hay una reclamación en proceso
    dato_id = request.session.get('reclamacion_dato_id')
    email = request.session.get('reclamacion_email')
    
    if not dato_id or not email:
        messages.error(request, 'No hay un proceso de reclamación activo.')
        return redirect('datos_archivados:buscar_usuario')
    
    try:
        # Obtener el dato archivado
        dato = DatoArchivadoDinamico.objects.get(id=dato_id)
        datos = dato.datos_originales
        username = datos.get('username') or datos.get('user') or 'Usuario'
        
        # Generar nuevo código de 4 dígitos
        codigo = ''.join(random.choices(string.digits, k=4))
        
        # Eliminar códigos anteriores para este email
        CodigoVerificacionReclamacion.objects.filter(email=email).delete()
        
        # Crear nuevo código
        fecha_expiracion = timezone.now() + timedelta(minutes=15)
        CodigoVerificacionReclamacion.objects.create(
            email=email,
            codigo=codigo,
            dato_archivado=dato,
            fecha_expiracion=fecha_expiracion
        )
        
        # Enviar email
        email_text = f'''Hola {username},

Se ha solicitado un nuevo código de verificación para reclamar su cuenta archivada en el Centro Fray Bartolomé de las Casas.

Su nuevo código de verificación es:

{codigo}

Este código expirará en 15 minutos.

Si usted no solicitó esta acción, ignore este mensaje.

Centro Fray Bartolomé de las Casas'''
        
        send_mail(
            'Nuevo Código de Verificación - Reclamación de Cuenta',
            email_text,
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=False,
        )
        
        messages.success(request, f'Se ha enviado un nuevo código de verificación a {email}')
        return redirect('datos_archivados:verificar_codigo_reclamacion')
        
    except Exception as e:
        messages.error(request, f'Error al reenviar el código: {str(e)}')
        return redirect('datos_archivados:verificar_codigo_reclamacion')

def verificar_codigo_reclamacion(request):
    """Vista para verificar el código de 4 dígitos"""
    from .models import CodigoVerificacionReclamacion
    
    # Verificar que hay una reclamación en proceso
    dato_id = request.session.get('reclamacion_dato_id')
    email = request.session.get('reclamacion_email')
    
    if not dato_id or not email:
        messages.error(request, 'No hay un proceso de reclamación activo.')
        return redirect('datos_archivados:buscar_usuario')
    
    if request.method == 'POST':
        codigo_ingresado = request.POST.get('code', '').strip()
        
        if not codigo_ingresado:
            messages.error(request, 'Debe ingresar el código de verificación.')
            return render(request, 'datos_archivados/verificar_codigo_reclamacion.html', {
                'email': email
            })
        
        try:
            # Buscar código válido (el más reciente si hay múltiples)
            codigo_verificacion = CodigoVerificacionReclamacion.objects.filter(
                email=email,
                codigo=codigo_ingresado,
                usado=False
            ).order_by('-fecha_creacion').first()
            
            if not codigo_verificacion:
                raise CodigoVerificacionReclamacion.DoesNotExist()
            
            if codigo_verificacion.is_expired():
                messages.error(request, 'El código de verificación ha expirado. Solicite uno nuevo.')
                return render(request, 'datos_archivados/verificar_codigo_reclamacion.html', {
                    'email': email
                })
            
            # Marcar código como usado SOLO después de verificar que es válido
            codigo_verificacion.usado = True
            codigo_verificacion.save()
            
            # Redirigir a la página de reclamación
            messages.success(request, 'Código verificado correctamente.')
            return redirect('datos_archivados:reclamar_usuario_dinamico', dato_id=dato_id)
            
        except CodigoVerificacionReclamacion.DoesNotExist:
            messages.error(request, 'Código de verificación incorrecto.')
            return render(request, 'datos_archivados/verificar_codigo_reclamacion.html', {
                'email': email
            })
    
    return render(request, 'datos_archivados/verificar_codigo_reclamacion.html', {
        'email': email
    })

def reclamar_usuario_dinamico(request, dato_id):
    """Vista para reclamar un usuario desde datos archivados dinámicos (después de verificación)"""
    from .models import DatoArchivadoDinamico
    from django.contrib.auth.models import User, Group
    from django.contrib.auth import login
    from django.shortcuts import get_object_or_404
    from django.core.mail import send_mail
    from django.conf import settings
    
    # Verificar que el proceso de verificación se completó
    session_dato_id = request.session.get('reclamacion_dato_id')
    if not session_dato_id or int(session_dato_id) != dato_id:
        messages.error(request, 'Debe completar la verificación por email primero.')
        return redirect('datos_archivados:buscar_usuario')
    
    # Obtener el dato archivado
    dato = get_object_or_404(DatoArchivadoDinamico, id=dato_id)
    datos = dato.datos_originales
    
    if request.method == 'POST':
        # Obtener datos del formulario
        nueva_password = request.POST.get('nueva_password')
        confirmar_password = request.POST.get('confirmar_password')
        
        # Validaciones
        if not nueva_password or not confirmar_password:
            messages.error(request, 'Debe proporcionar una nueva contraseña.')
            return render(request, 'datos_archivados/reclamar_usuario_dinamico.html', {
                'dato': dato,
                'datos': datos
            })
        
        if nueva_password != confirmar_password:
            messages.error(request, 'Las contraseñas no coinciden.')
            return render(request, 'datos_archivados/reclamar_usuario_dinamico.html', {
                'dato': dato,
                'datos': datos
            })
        
        if len(nueva_password) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            return render(request, 'datos_archivados/reclamar_usuario_dinamico.html', {
                'dato': dato,
                'datos': datos
            })
        
        # Extraer información del usuario
        username_original = datos.get('username') or datos.get('user') or f"usuario_{dato.id_original}"
        email = datos.get('email') or datos.get('correo') or ''
        first_name = datos.get('first_name') or datos.get('nombre') or datos.get('name') or ''
        last_name = datos.get('last_name') or datos.get('apellido') or datos.get('apellidos') or ''
        
        # Generar username único con protección contra condiciones de carrera
        from django.db import IntegrityError
        
        user = None
        max_intentos = 100
        
        for intento in range(max_intentos):
            # Generar username candidato
            if intento == 0:
                username_candidato = username_original
            else:
                username_candidato = f"{username_original}{intento - 1}"
            
            # Verificar si ya existe
            if User.objects.filter(username=username_candidato).exists():
                continue
            
            # Intentar crear el usuario
            try:
                user = User.objects.create_user(
                    username=username_candidato,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    password=nueva_password,
                    is_active=True
                )
                username_final = username_candidato
                break  # Usuario creado exitosamente
            except IntegrityError:
                # El username fue creado por otro proceso entre la verificación y la creación
                # Continuar con el siguiente número
                continue
        
        if not user:
            messages.error(request, f'No se pudo generar un username único para {username_original}. Contacte al administrador.')
            return render(request, 'datos_archivados/reclamar_usuario_dinamico.html', {
                'dato': dato,
                'datos': datos
            })
        
        try:
            
            # Asignar al grupo Estudiantes por defecto
            try:
                grupo_estudiantes = Group.objects.get(name='Estudiantes')
                user.groups.add(grupo_estudiantes)
            except Group.DoesNotExist:
                # Crear el grupo si no existe
                grupo_estudiantes = Group.objects.create(name='Estudiantes')
                user.groups.add(grupo_estudiantes)
            
            # Enviar email de confirmación con detalles de la cuenta
            username_cambio = username_original != username_final
            
            if username_cambio:
                mensaje_username = f'''IMPORTANTE: Su nombre de usuario ha cambiado
Nombre de usuario original: {username_original}
Nuevo nombre de usuario: {username_final}

Esto se debe a que ya existía un usuario con el nombre "{username_original}" en el sistema.
Por favor, use "{username_final}" para futuros inicios de sesión.'''
            else:
                mensaje_username = f'Su nombre de usuario es: {username_final}'
            
            mensaje = f'''¡Bienvenido de vuelta al Centro Fray Bartolomé de las Casas!

Su cuenta ha sido reactivada exitosamente desde los datos archivados.

DETALLES DE SU CUENTA:
{mensaje_username}
Correo electrónico: {user.email}
Nombre completo: {user.get_full_name()}

Ahora puede acceder a todos los servicios del sistema usando sus credenciales.

Si tiene alguna pregunta o necesita ayuda, no dude en contactarnos.

Saludos cordiales,
Centro Fray Bartolomé de las Casas'''
            
            try:
                send_mail(
                    'Cuenta Reactivada - Centro Fray Bartolomé de las Casas',
                    mensaje,
                    settings.DEFAULT_FROM_EMAIL,
                    [user.email],
                    fail_silently=False,
                )
            except Exception as e:
                # No fallar si el email no se puede enviar
                pass
            
            # Limpiar sesión
            if 'reclamacion_dato_id' in request.session:
                del request.session['reclamacion_dato_id']
            if 'reclamacion_email' in request.session:
                del request.session['reclamacion_email']
            
            # Iniciar sesión automáticamente con backend específico
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            
            messages.success(
                request, 
                f'¡Bienvenido de vuelta, {user.get_full_name() or user.username}! '
                'Su cuenta ha sido reactivada exitosamente desde los datos archivados. '
                'Se le ha enviado un email con los detalles de su cuenta.'
            )
            
            return redirect('principal:login_redirect')
            
        except Exception as e:
            messages.error(request, f'Error al crear la cuenta: {str(e)}')
    
    context = {
        'dato': dato,
        'datos': datos,
        'username': datos.get('username') or datos.get('user') or f"usuario_{dato.id_original}",
        'email': datos.get('email') or datos.get('correo') or '',
        'first_name': datos.get('first_name') or datos.get('nombre') or datos.get('name') or '',
        'last_name': datos.get('last_name') or datos.get('apellido') or datos.get('apellidos') or '',
    }
    
    return render(request, 'datos_archivados/reclamar_usuario_dinamico.html', context)

@login_required
def buscar_datos_ajax(request):
    """Vista AJAX para búsqueda en tiempo real de datos archivados"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.db.models import Q, Count
        
        search_query = request.GET.get('q', '').strip()
        search_type = request.GET.get('type', 'global')
        limit = int(request.GET.get('limit', 10))
        
        if not search_query:
            return JsonResponse({
                'results': [],
                'total': 0,
                'query': search_query
            })
        
        # Construir query de búsqueda
        queryset = DatoArchivadoDinamico.objects.all()
        
        if search_type == 'tabla':
            queryset = queryset.filter(tabla_origen__icontains=search_query)
        elif search_type == 'contenido':
            queryset = queryset.filter(
                Q(datos_originales__icontains=search_query) |
                Q(nombre_registro__icontains=search_query)
            )
        else:  # global
            queryset = queryset.filter(
                Q(tabla_origen__icontains=search_query) |
                Q(datos_originales__icontains=search_query) |
                Q(nombre_registro__icontains=search_query)
            )
        
        # Obtener resultados agrupados por tabla
        tablas_resultados = queryset.values('tabla_origen').annotate(
            total=Count('id')
        ).order_by('-total', 'tabla_origen')[:limit]
        
        # Obtener algunos registros de ejemplo para cada tabla
        resultados = []
        for tabla in tablas_resultados:
            tabla_nombre = tabla['tabla_origen']
            total_registros = tabla['total']
            
            # Obtener algunos registros de ejemplo
            ejemplos = queryset.filter(tabla_origen=tabla_nombre)[:3]
            ejemplos_data = []
            
            for ejemplo in ejemplos:
                # Extraer información relevante del registro
                datos = ejemplo.datos_originales
                nombre = ejemplo.obtener_nombre_legible()
                
                # Buscar campos que contengan el término de búsqueda
                campos_coincidentes = []
                if search_type in ['contenido', 'global']:
                    for campo, valor in datos.items():
                        if search_query.lower() in str(valor).lower():
                            campos_coincidentes.append({
                                'campo': campo,
                                'valor': str(valor)[:100] + ('...' if len(str(valor)) > 100 else '')
                            })
                
                ejemplos_data.append({
                    'id': ejemplo.id,
                    'nombre': nombre,
                    'campos_coincidentes': campos_coincidentes[:3]  # Máximo 3 campos
                })
            
            resultados.append({
                'tabla': tabla_nombre,
                'total_registros': total_registros,
                'ejemplos': ejemplos_data
            })
        
        return JsonResponse({
            'results': resultados,
            'total': queryset.count(),
            'query': search_query,
            'search_type': search_type
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'Error en búsqueda: {str(e)}',
            'results': [],
            'total': 0
        })

@login_required
def buscar_en_tabla_ajax(request, tabla):
    """Vista AJAX para búsqueda en tiempo real dentro de una tabla específica"""
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.db.models import Q
        
        search_query = request.GET.get('q', '').strip()
        order_by = request.GET.get('order_by', 'fecha_migracion')
        order_direction = request.GET.get('order_direction', 'desc')
        limit = int(request.GET.get('limit', 20))
        
        # Query base para la tabla específica
        queryset = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla)
        
        # Aplicar filtro de búsqueda si existe
        if search_query:
            queryset = queryset.filter(
                Q(datos_originales__icontains=search_query) |
                Q(nombre_registro__icontains=search_query)
            )
        
        # Aplicar ordenamiento
        valid_order_fields = {
            'fecha_migracion': 'fecha_migracion',
            'id_original': 'id_original',
            'nombre': 'nombre_registro',
            'tabla': 'tabla_origen'
        }
        
        if order_by in valid_order_fields:
            order_field = valid_order_fields[order_by]
            if order_direction == 'desc':
                order_field = f'-{order_field}'
            queryset = queryset.order_by(order_field)
        
        # Obtener resultados limitados
        resultados = queryset[:limit]
        
        # Para ordenamiento por nombre, necesitamos ordenar después
        if order_by == 'nombre':
            resultados_list = list(resultados)
            resultados_list.sort(
                key=lambda x: (x.obtener_nombre_legible() or '').lower(),
                reverse=(order_direction == 'desc')
            )
            resultados = resultados_list
        
        # Preparar datos para JSON
        datos_json = []
        for dato in resultados:
            # Resaltar términos de búsqueda en el nombre
            nombre = dato.obtener_nombre_legible()
            if search_query and nombre:
                # Simple highlighting (se puede mejorar)
                nombre_resaltado = nombre.replace(
                    search_query, 
                    f'<mark>{search_query}</mark>'
                )
            else:
                nombre_resaltado = nombre
            
            # Buscar campos que coincidan con la búsqueda
            campos_coincidentes = []
            if search_query:
                for campo, valor in dato.datos_originales.items():
                    if search_query.lower() in str(valor).lower():
                        valor_str = str(valor)
                        if len(valor_str) > 100:
                            valor_str = valor_str[:100] + '...'
                        campos_coincidentes.append({
                            'campo': campo,
                            'valor': valor_str
                        })
                        if len(campos_coincidentes) >= 3:  # Máximo 3 campos
                            break
            
            datos_json.append({
                'id': dato.pk,
                'id_original': dato.id_original,
                'nombre': nombre,
                'nombre_resaltado': nombre_resaltado,
                'fecha_migracion': dato.fecha_migracion.strftime('%d/%m/%Y %H:%M'),
                'campos_coincidentes': campos_coincidentes,
                'url_detalle': f'/datos-archivados/datos/{dato.pk}/'
            })
        
        return JsonResponse({
            'success': True,
            'resultados': datos_json,
            'total_encontrados': queryset.count(),
            'total_mostrados': len(datos_json),
            'query': search_query,
            'tabla': tabla,
            'order_by': order_by,
            'order_direction': order_direction
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en búsqueda: {str(e)}',
            'resultados': []
        })

@login_required
def debug_permisos(request):
    """Vista para diagnosticar permisos del usuario"""
    user = request.user
    
    # Información del usuario
    info = f"""
    <h2>Información de Debug - Usuario: {user.username}</h2>
    <p><strong>Autenticado:</strong> {user.is_authenticated}</p>
    <p><strong>Es superusuario:</strong> {user.is_superuser}</p>
    <p><strong>Es staff:</strong> {user.is_staff}</p>
    <p><strong>Grupos del usuario:</strong></p>
    <ul>
    """
    
    for grupo in user.groups.all():
        info += f"<li>{grupo.name}</li>"
    
    info += "</ul>"
    
    # Verificar específicamente el grupo Secretaria
    es_secretaria_check = user.groups.filter(name='Secretaria').exists()
    info += f"<p><strong>¿Es Secretaria?:</strong> {es_secretaria_check}</p>"
    
    # Información adicional
    info += f"""
    <p><strong>ID del usuario:</strong> {user.id}</p>
    <p><strong>Email:</strong> {user.email}</p>
    <p><strong>Nombre completo:</strong> {user.get_full_name()}</p>
    
    <hr>
    <h3>URLs de prueba:</h3>
    <p><a href="/datos-archivados/">Dashboard Datos Archivados</a></p>
    <p><a href="/datos-archivados/migracion/configurar/">Configurar Migración</a></p>
    <p><a href="/admin/">Admin Django</a></p>
    """
    
    return HttpResponse(info)


def agregar_campos_faltantes_a_modelo(modelo, campos_necesarios, logger):
    """
    Agrega campos faltantes a un modelo de Django dinámicamente usando ALTER TABLE
    """
    from django.db import connection
    from datetime import datetime, date
    
    campos_actuales = {f.name for f in modelo._meta.get_fields()}
    campos_agregados = []
    
    with connection.cursor() as cursor:
        tabla_nombre = modelo._meta.db_table
        
        for campo_nombre, valor_ejemplo in campos_necesarios.items():
            if campo_nombre not in campos_actuales and campo_nombre not in ['id', 'pk']:
                try:
                    # Determinar tipo de campo basado en el valor
                    if isinstance(valor_ejemplo, bool):
                        tipo_sql = 'BOOLEAN DEFAULT FALSE'
                    elif isinstance(valor_ejemplo, int):
                        tipo_sql = 'INTEGER NULL'
                    elif isinstance(valor_ejemplo, float):
                        tipo_sql = 'DOUBLE PRECISION NULL'
                    elif isinstance(valor_ejemplo, (datetime, date)):
                        tipo_sql = 'TIMESTAMP NULL'
                    elif isinstance(valor_ejemplo, str) and len(valor_ejemplo) > 255:
                        tipo_sql = 'TEXT NULL'
                    else:
                        # Por defecto, VARCHAR
                        tipo_sql = 'VARCHAR(255) NULL'
                    
                    # Ejecutar ALTER TABLE
                    sql = f'ALTER TABLE {tabla_nombre} ADD COLUMN IF NOT EXISTS "{campo_nombre}" {tipo_sql}'
                    cursor.execute(sql)
                    
                    campos_agregados.append(campo_nombre)
                    logger.info(f"✅ Campo agregado: {tabla_nombre}.{campo_nombre} ({tipo_sql})")
                    
                except Exception as e:
                    logger.warning(f"⚠️ No se pudo agregar campo {campo_nombre} a {tabla_nombre}: {str(e)}")
    
    return campos_agregados

def mapear_campos_ingles_espanol(datos_origen, logger=None):
    """
    Mapea campos de inglés a español y mantiene ambos
    Retorna un diccionario con campos mapeados + campos originales
    """
    # MAPEO DE CAMPOS: Traducir nombres de inglés a español
    mapeo_campos = {
        'nacionality': 'nacionalidad',
        'numberidentification': 'carnet',
        'phone': 'telephone',
        'cellphone': 'movil',
        'street': 'address',
        'city': 'location',
        'state': 'provincia',
        'degree': 'grado',
        'ocupation': 'ocupacion',
        'title': 'titulo',
        'gender': 'sexo',
        'name': 'first_name',
        'lastname': 'last_name',
        'email': 'email',  # Mantener email
        'photo': 'image',  # Mapear photo a image
        'resume': 'titulo',  # Resume puede ser título
    }
    
    # Crear copia con todos los datos originales
    datos_mapeados = datos_origen.copy()
    campos_mapeados_count = 0
    
    # Aplicar mapeo: agregar campos traducidos
    for campo_origen, campo_destino in mapeo_campos.items():
        if campo_origen in datos_origen:
            datos_mapeados[campo_destino] = datos_origen[campo_origen]
            campos_mapeados_count += 1
            if logger:
                logger.info(f"🔄 Mapeando: {campo_origen} → {campo_destino} = {datos_origen[campo_origen]}")
    
    if logger and campos_mapeados_count > 0:
        logger.info(f"📋 Total de campos mapeados: {campos_mapeados_count}")
        logger.info(f"📋 Total de campos disponibles: {len(datos_mapeados)} (originales + mapeados)")
    
    return datos_mapeados

def copiar_todos_los_campos(objeto_destino, datos_origen, campos_excluir=None, logger=None):
    """
    Copia TODOS los campos de datos_origen al objeto_destino
    NOTA: Los campos deben existir previamente (creados en la fase de detección)
    """
    if campos_excluir is None:
        campos_excluir = ['id', 'pk']
    
    campos_copiados = 0
    campos_no_encontrados = []
    
    # Obtener los nombres de campos del modelo
    campos_modelo = {f.name for f in objeto_destino._meta.get_fields() if not f.many_to_many and not f.one_to_many}
    
    for campo_nombre, valor in datos_origen.items():
        if campo_nombre in campos_excluir:
            continue
        
        # Verificar si el campo existe en el modelo
        if campo_nombre not in campos_modelo:
            campos_no_encontrados.append(campo_nombre)
            if logger:
                logger.debug(f"⚠️ Campo '{campo_nombre}' no existe en el modelo {objeto_destino.__class__.__name__}")
            continue
        
        try:
            # Establecer el valor
            setattr(objeto_destino, campo_nombre, valor)
            campos_copiados += 1
            if logger:
                logger.debug(f"✅ Campo copiado: {campo_nombre} = {valor}")
        except Exception as e:
            if logger:
                logger.warning(f"❌ No se pudo copiar campo {campo_nombre}: {str(e)}")
    
    if logger and campos_copiados > 0:
        logger.info(f"📊 Total campos copiados: {campos_copiados}")
        if campos_no_encontrados:
            logger.debug(f"⚠️ Campos no encontrados: {', '.join(campos_no_encontrados[:5])}")
    
    return campos_copiados

@login_required
def combinar_datos_archivados(request):
    """Vista para combinar TODAS las tablas archivadas con las tablas activas de la base de datos
    
    Esta función:
    1. Detecta campos faltantes en las tablas actuales
    2. Crea esos campos automáticamente usando ALTER TABLE
    3. Copia TODA la información de las tablas archivadas
    """
    if not tiene_permisos_datos_archivados(request.user):
        return JsonResponse({'success': False, 'error': 'Sin permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        from .models import DatoArchivadoDinamico
        from django.contrib.auth.models import User
        from accounts.models import Registro
        from principal.models import (
            CursoAcademico, Curso, Matriculas, 
            Asistencia, Calificaciones, NotaIndividual
        )
        from django.db import transaction, IntegrityError
        from django.db.models import Q
        from datetime import datetime, date
        from decimal import Decimal
        import logging
        
        logger = logging.getLogger(__name__)
        
        # Contadores
        estadisticas = {
            'usuarios_combinados': 0,
            'registros_combinados': 0,
            'cursos_academicos_combinados': 0,
            'cursos_combinados': 0,
            'matriculas_combinadas': 0,
            'asistencias_combinadas': 0,
            'calificaciones_combinadas': 0,
            'notas_combinadas': 0,
            'otras_tablas': 0
        }
        campos_agregados = []
        errores = []
        
        # Mapeo de usuarios archivados a usuarios actuales
        mapeo_usuarios = {}
        
        with transaction.atomic():
            # 0. DETECTAR Y AGREGAR CAMPOS FALTANTES
            logger.info("=== Detectando campos faltantes en modelos ===")
            
            # Analizar TODOS los campos necesarios para User
            logger.info("=== Analizando TODOS los campos de usuarios ===")
            datos_users_all = DatoArchivadoDinamico.objects.filter(tabla_origen='auth_user')
            
            # Recopilar TODOS los campos únicos de TODOS los usuarios
            todos_campos_user = {}
            for dato_user in datos_users_all:
                for campo, valor in dato_user.datos_originales.items():
                    if campo not in todos_campos_user:
                        todos_campos_user[campo] = valor
            
            logger.info(f"📊 Total de campos únicos encontrados en usuarios: {len(todos_campos_user)}")
            logger.info(f"📋 Campos: {list(todos_campos_user.keys())}")
            
            # Crear TODOS los campos necesarios
            if todos_campos_user:
                campos_agregados_user = agregar_campos_faltantes_a_modelo(User, todos_campos_user, logger)
                if campos_agregados_user:
                    campos_agregados.extend([f"User.{c}" for c in campos_agregados_user])
                    logger.info(f"✅ Total de campos agregados a User: {len(campos_agregados_user)}")
            
            # Analizar TODOS los campos necesarios para Registro (Estudiantes y Profesores)
            logger.info("=== Analizando TODOS los campos de registros ===")
            datos_registros_all = DatoArchivadoDinamico.objects.filter(
                Q(tabla_origen='Docencia_studentpersonalinformation') |
                Q(tabla_origen='Docencia_teacherpersonalinformation') |
                Q(tabla_origen='accounts_registro')
            )
            
            # Recopilar TODOS los campos únicos de TODOS los registros
            todos_campos_registro = {}
            for dato_reg in datos_registros_all:
                for campo, valor in dato_reg.datos_originales.items():
                    if campo not in todos_campos_registro:
                        todos_campos_registro[campo] = valor
            
            logger.info(f"📊 Total de campos únicos encontrados en registros: {len(todos_campos_registro)}")
            logger.info(f"📋 Campos: {list(todos_campos_registro.keys())}")
            
            # Crear TODOS los campos necesarios
            if todos_campos_registro:
                campos_agregados_registro = agregar_campos_faltantes_a_modelo(Registro, todos_campos_registro, logger)
                if campos_agregados_registro:
                    campos_agregados.extend([f"Registro.{c}" for c in campos_agregados_registro])
                    logger.info(f"✅ Total de campos agregados a Registro: {len(campos_agregados_registro)}")
            
            # 1. COMBINAR auth_user - COPIAR TODOS LOS CAMPOS
            logger.info("=== Iniciando combinación de auth_user (TODOS LOS CAMPOS) ===")
            datos_auth_user = DatoArchivadoDinamico.objects.filter(tabla_origen='auth_user')
            
            for dato in datos_auth_user:
                try:
                    datos = dato.datos_originales
                    username = datos.get('username', '')
                    email = datos.get('email', '')
                    id_original = dato.id_original
                    
                    if not username:
                        continue
                    
                    # Buscar si el usuario ya existe
                    usuario_existente = User.objects.filter(
                        Q(username=username) | (Q(email=email) if email else Q(pk=None))
                    ).first()
                    
                    if usuario_existente:
                        # COPIAR TODOS LOS CAMPOS del usuario archivado
                        campos_copiados = copiar_todos_los_campos(
                            usuario_existente, 
                            datos, 
                            campos_excluir=['id', 'pk', 'username'],  # No cambiar username
                            logger=logger
                        )
                        
                        # IMPORTANTE: Procesar contraseña (hasheada o texto plano)
                        password_original = datos.get('password')
                        if password_original:
                            # Verificar si la contraseña ya está hasheada
                            if password_original.startswith(('pbkdf2_sha256$', 'bcrypt$', 'argon2$', 'sha1$', 'md5$')):
                                # Ya está hasheada, copiar directamente
                                usuario_existente.password = password_original
                                logger.info(f"✅ Contraseña hasheada copiada para usuario: {username}")
                            else:
                                # Está en texto plano, hashear antes de guardar
                                usuario_existente.set_password(password_original)
                                logger.info(f"✅ Contraseña en texto plano hasheada para usuario: {username}")
                        
                        usuario_existente.save()
                        mapeo_usuarios[id_original] = usuario_existente
                        estadisticas['usuarios_combinados'] += 1
                        logger.info(f"✅ Usuario actualizado: {username} ({campos_copiados} campos copiados)")
                    else:
                        # Crear nuevo usuario CON TODOS LOS CAMPOS
                        # Crear usuario base
                        nuevo_usuario = User(username=username)
                        
                        # COPIAR TODOS LOS CAMPOS automáticamente
                        campos_copiados = copiar_todos_los_campos(
                            nuevo_usuario,
                            datos,
                            campos_excluir=['id', 'pk'],
                            logger=logger
                        )
                        
                        # IMPORTANTE: Procesar contraseña (hasheada o texto plano)
                        password_original = datos.get('password')
                        if password_original:
                            # Verificar si la contraseña ya está hasheada
                            if password_original.startswith(('pbkdf2_sha256$', 'bcrypt$', 'argon2$', 'sha1$', 'md5$')):
                                # Ya está hasheada, copiar directamente
                                nuevo_usuario.password = password_original
                                logger.info(f"✅ Contraseña hasheada copiada para nuevo usuario: {username}")
                            else:
                                # Está en texto plano, hashear antes de guardar
                                nuevo_usuario.set_password(password_original)
                                logger.info(f"✅ Contraseña en texto plano hasheada para nuevo usuario: {username}")
                        else:
                            # Si no hay contraseña, establecer una no utilizable
                            nuevo_usuario.set_unusable_password()
                            logger.warning(f"⚠️ No se encontró contraseña para usuario: {username}")
                        
                        # Convertir fechas si son strings
                        if hasattr(nuevo_usuario, 'date_joined') and isinstance(nuevo_usuario.date_joined, str):
                            try:
                                nuevo_usuario.date_joined = datetime.fromisoformat(
                                    nuevo_usuario.date_joined.replace('Z', '+00:00')
                                )
                            except:
                                nuevo_usuario.date_joined = datetime.now()
                        
                        if hasattr(nuevo_usuario, 'last_login') and isinstance(nuevo_usuario.last_login, str):
                            try:
                                nuevo_usuario.last_login = datetime.fromisoformat(
                                    nuevo_usuario.last_login.replace('Z', '+00:00')
                                )
                            except:
                                nuevo_usuario.last_login = None
                        
                        # Guardar el usuario
                        try:
                            nuevo_usuario.save()
                            mapeo_usuarios[id_original] = nuevo_usuario
                            estadisticas['usuarios_combinados'] += 1
                            logger.info(f"✅ Usuario creado: {username} ({campos_copiados} campos copiados)")
                        except IntegrityError as e:
                            # Si hay error de duplicado, intentar buscar el usuario existente
                            logger.warning(f"⚠️ Usuario {username} ya existe, buscando para actualizar...")
                            usuario_duplicado = User.objects.filter(username=username).first()
                            if usuario_duplicado:
                                mapeo_usuarios[id_original] = usuario_duplicado
                                logger.info(f"✅ Usuario duplicado encontrado y mapeado: {username}")
                            else:
                                logger.error(f"❌ No se pudo crear ni encontrar usuario: {username}")
                                continue
                            
                except Exception as e:
                    error_msg = f"Error procesando usuario {dato.id_original}: {str(e)}"
                    logger.error(error_msg)
                    errores.append(error_msg)
                    continue
            
            # 1.5. ASIGNAR GRUPOS A USUARIOS (auth_user_groups)
            logger.info("=== Asignando grupos a usuarios ===")
            from django.contrib.auth.models import Group
            
            datos_user_groups = DatoArchivadoDinamico.objects.filter(
                tabla_origen='auth_user_groups'
            )
            
            for dato in datos_user_groups:
                try:
                    datos = dato.datos_originales
                    user_id_original = datos.get('user_id')
                    group_id_original = datos.get('group_id')
                    
                    if not user_id_original or not group_id_original:
                        continue
                    
                    # Buscar el usuario en el mapeo
                    usuario = mapeo_usuarios.get(user_id_original)
                    
                    if not usuario:
                        logger.warning(f"No se encontró usuario para asignar grupo: user_id={user_id_original}")
                        continue
                    
                    # Buscar el grupo en los datos archivados
                    dato_grupo = DatoArchivadoDinamico.objects.filter(
                        tabla_origen='auth_group',
                        id_original=group_id_original
                    ).first()
                    
                    if dato_grupo:
                        nombre_grupo = dato_grupo.datos_originales.get('name')
                        
                        if nombre_grupo:
                            # Buscar o crear el grupo
                            grupo, created = Group.objects.get_or_create(name=nombre_grupo)
                            
                            # Asignar el usuario al grupo
                            if not usuario.groups.filter(id=grupo.id).exists():
                                usuario.groups.add(grupo)
                                logger.info(f"Usuario {usuario.username} agregado al grupo {nombre_grupo}")
                    
                except Exception as e:
                    error_msg = f"Error asignando grupo: {str(e)}"
                    logger.error(error_msg)
                    errores.append(error_msg)
                    continue
            
            # 2. IDENTIFICAR PROFESORES PRIMERO (para no agregarlos a Estudiantes)
            logger.info("=== Identificando profesores ===")
            user_ids_profesores = set()
            datos_teacher_info_temp = DatoArchivadoDinamico.objects.filter(
                tabla_origen='Docencia_teacherpersonalinformation'
            )
            for dato_temp in datos_teacher_info_temp:
                user_id = dato_temp.datos_originales.get('user_id')
                if user_id:
                    user_ids_profesores.add(user_id)
            logger.info(f"📋 Total de user_ids identificados como profesores: {len(user_ids_profesores)}")
            logger.info(f"📋 IDs de profesores: {user_ids_profesores}")
            
            # 2. COMBINAR Docencia_studentpersonalinformation con accounts_registro - TODOS LOS CAMPOS
            logger.info("=== Iniciando combinación de Docencia_studentpersonalinformation (TODOS LOS CAMPOS) ===")
            datos_student_info = DatoArchivadoDinamico.objects.filter(
                Q(tabla_origen='Docencia_studentpersonalinformation') |
                Q(tabla_origen='accounts_registro')
            )
            
            for dato in datos_student_info:
                try:
                    datos = dato.datos_originales
                    user_id_original = datos.get('user_id')
                    
                    if not user_id_original:
                        continue
                    
                    # SALTAR si este usuario es un profesor
                    if user_id_original in user_ids_profesores:
                        logger.info(f"⏭️ Saltando user_id {user_id_original} porque es profesor")
                        continue
                    
                    # Buscar usuario en el mapeo
                    usuario = mapeo_usuarios.get(user_id_original)
                    
                    if not usuario:
                        # Buscar en datos archivados
                        dato_usuario = DatoArchivadoDinamico.objects.filter(
                            tabla_origen='auth_user',
                            id_original=user_id_original
                        ).first()
                        
                        if dato_usuario:
                            username = dato_usuario.datos_originales.get('username')
                            email = dato_usuario.datos_originales.get('email')
                            
                            if username:
                                usuario = User.objects.filter(username=username).first()
                            if not usuario and email:
                                usuario = User.objects.filter(email=email).first()
                    
                    if not usuario:
                        logger.warning(f"No se encontró usuario para registro {dato.id_original}")
                        continue
                    
                    # Buscar o crear registro
                    registro, created = Registro.objects.get_or_create(user=usuario)
                    
                    # LOG: Mostrar datos que se van a copiar
                    logger.info(f"📝 Datos a copiar para usuario {usuario.username}:")
                    logger.info(f"   Campos disponibles: {list(datos.keys())}")
                    
                    # MAPEAR campos de inglés a español
                    datos_mapeados = mapear_campos_ingles_espanol(datos, logger)
                    
                    # COPIAR TODOS LOS CAMPOS automáticamente (con mapeo aplicado)
                    campos_copiados = copiar_todos_los_campos(
                        registro,
                        datos_mapeados,
                        campos_excluir=['id', 'pk', 'user_id', 'user'],
                        logger=logger
                    )
                    
                    # Asegurar que el user esté establecido
                    registro.user = usuario
                    
                    # LOG: Mostrar valores después de copiar
                    logger.info(f"📊 Valores copiados al registro:")
                    logger.info(f"   nacionalidad: {registro.nacionalidad}")
                    logger.info(f"   carnet: {registro.carnet}")
                    logger.info(f"   telephone: {registro.telephone}")
                    logger.info(f"   address: {registro.address}")
                    
                    # Guardar el registro
                    try:
                        registro.save()
                        estadisticas['registros_combinados'] += 1
                        logger.info(f"✅ Registro {'creado' if created else 'actualizado'} para usuario {usuario.username} ({campos_copiados} campos copiados)")
                        
                        # Verificar que se guardó correctamente
                        registro_verificado = Registro.objects.get(user=usuario)
                        logger.info(f"🔍 Verificación - nacionalidad guardada: {registro_verificado.nacionalidad}")
                        logger.info(f"🔍 Verificación - carnet guardado: {registro_verificado.carnet}")
                        logger.info(f"🔍 Verificación - telephone guardado: {registro_verificado.telephone}")
                    except Exception as save_error:
                        logger.error(f"❌ Error al guardar registro: {str(save_error)}")
                        raise
                        
                except Exception as e:
                    error_msg = f"Error procesando registro {dato.id_original}: {str(e)}"
                    logger.error(error_msg)
                    import traceback
                    logger.error(f"Traceback: {traceback.format_exc()}")
                    errores.append(error_msg)
                    continue
            
            # 2.5. COMBINAR Docencia_teacherpersonalinformation con accounts_registro - TODOS LOS CAMPOS
            logger.info("=== Iniciando combinación de Docencia_teacherpersonalinformation (TODOS LOS CAMPOS) ===")
            datos_teacher_info = DatoArchivadoDinamico.objects.filter(
                tabla_origen='Docencia_teacherpersonalinformation'
            )
            
            # Obtener o crear el grupo Profesores
            grupo_profesores, _ = Group.objects.get_or_create(name='Profesores')
            
            for dato in datos_teacher_info:
                try:
                    datos = dato.datos_originales
                    user_id_original = datos.get('user_id')
                    
                    if not user_id_original:
                        continue
                    
                    # Buscar usuario en el mapeo
                    usuario = mapeo_usuarios.get(user_id_original)
                    
                    if not usuario:
                        # Buscar en datos archivados
                        dato_usuario = DatoArchivadoDinamico.objects.filter(
                            tabla_origen='auth_user',
                            id_original=user_id_original
                        ).first()
                        
                        if dato_usuario:
                            username = dato_usuario.datos_originales.get('username')
                            email = dato_usuario.datos_originales.get('email')
                            
                            if username:
                                usuario = User.objects.filter(username=username).first()
                            if not usuario and email:
                                usuario = User.objects.filter(email=email).first()
                    
                    if not usuario:
                        logger.warning(f"No se encontró usuario para registro de profesor {dato.id_original}")
                        continue
                    
                    # ASIGNAR AL GRUPO PROFESORES y QUITAR de Estudiantes si está
                    if not usuario.groups.filter(id=grupo_profesores.id).exists():
                        usuario.groups.add(grupo_profesores)
                        logger.info(f"✅ Usuario {usuario.username} agregado al grupo Profesores")
                    
                    # QUITAR del grupo Estudiantes si está
                    try:
                        grupo_estudiantes = Group.objects.get(name='Estudiantes')
                        if usuario.groups.filter(id=grupo_estudiantes.id).exists():
                            usuario.groups.remove(grupo_estudiantes)
                            logger.info(f"🔄 Usuario {usuario.username} removido del grupo Estudiantes")
                    except Group.DoesNotExist:
                        pass
                    
                    # Buscar o crear registro
                    registro, created = Registro.objects.get_or_create(user=usuario)
                    
                    # LOG: Mostrar datos que se van a copiar
                    logger.info(f"📝 Datos a copiar para PROFESOR {usuario.username}:")
                    logger.info(f"   Campos disponibles: {list(datos.keys())}")
                    
                    # MAPEAR campos de inglés a español
                    datos_mapeados = mapear_campos_ingles_espanol(datos, logger)
                    
                    # COPIAR TODOS LOS CAMPOS automáticamente (con mapeo aplicado)
                    campos_copiados = copiar_todos_los_campos(
                        registro,
                        datos_mapeados,
                        campos_excluir=['id', 'pk', 'user_id', 'user'],
                        logger=logger
                    )
                    
                    # Asegurar que el user esté establecido
                    registro.user = usuario
                    
                    # LOG: Mostrar valores después de copiar
                    logger.info(f"📊 Valores copiados al registro de PROFESOR:")
                    logger.info(f"   nacionalidad: {registro.nacionalidad}")
                    logger.info(f"   carnet: {registro.carnet}")
                    logger.info(f"   telephone: {registro.telephone}")
                    logger.info(f"   address: {registro.address}")
                    
                    # Guardar el registro
                    registro.save()
                    estadisticas['registros_combinados'] += 1
                    logger.info(f"✅ Registro de PROFESOR {'creado' if created else 'actualizado'} para usuario {usuario.username} ({campos_copiados} campos copiados)")
                        
                except Exception as e:
                    error_msg = f"Error procesando registro de profesor {dato.id_original}: {str(e)}"
                    logger.error(error_msg)
                    errores.append(error_msg)
                    continue
            
            # 3. COMBINAR principal_cursoacademico
            logger.info("=== Iniciando combinación de principal_cursoacademico ===")
            datos_cursos_academicos = DatoArchivadoDinamico.objects.filter(
                tabla_origen='principal_cursoacademico'
            )
            mapeo_cursos_academicos = {}
            
            for dato in datos_cursos_academicos:
                try:
                    datos = dato.datos_originales
                    nombre = datos.get('nombre', '')
                    
                    if not nombre:
                        continue
                    
                    # Buscar o crear curso académico
                    curso_academico, created = CursoAcademico.objects.get_or_create(
                        nombre=nombre,
                        defaults={
                            'activo': datos.get('activo', False),
                            'archivado': datos.get('archivado', True)
                        }
                    )
                    
                    if not created:
                        # Actualizar si existe
                        if datos.get('activo') is not None:
                            curso_academico.activo = datos.get('activo')
                        curso_academico.save()
                    
                    mapeo_cursos_academicos[dato.id_original] = curso_academico
                    estadisticas['cursos_academicos_combinados'] += 1
                    logger.info(f"Curso académico {'creado' if created else 'actualizado'}: {nombre}")
                    
                except Exception as e:
                    error_msg = f"Error procesando curso académico {dato.id_original}: {str(e)}"
                    logger.error(error_msg)
                    errores.append(error_msg)
                    continue
            
            # 4. COMBINAR principal_curso
            logger.info("=== Iniciando combinación de principal_curso ===")
            datos_cursos = DatoArchivadoDinamico.objects.filter(
                tabla_origen='principal_curso'
            )
            mapeo_cursos = {}
            
            for dato in datos_cursos:
                try:
                    datos = dato.datos_originales
                    name = datos.get('name', '')
                    teacher_id = datos.get('teacher_id')
                    curso_academico_id = datos.get('curso_academico_id')
                    
                    if not name:
                        continue
                    
                    # Buscar profesor
                    teacher = mapeo_usuarios.get(teacher_id)
                    if not teacher and teacher_id:
                        # Buscar en datos archivados
                        dato_teacher = DatoArchivadoDinamico.objects.filter(
                            tabla_origen='auth_user',
                            id_original=teacher_id
                        ).first()
                        if dato_teacher:
                            username = dato_teacher.datos_originales.get('username')
                            if username:
                                teacher = User.objects.filter(username=username).first()
                    
                    # Buscar curso académico
                    curso_academico = mapeo_cursos_academicos.get(curso_academico_id)
                    if not curso_academico and curso_academico_id:
                        dato_ca = DatoArchivadoDinamico.objects.filter(
                            tabla_origen='principal_cursoacademico',
                            id_original=curso_academico_id
                        ).first()
                        if dato_ca:
                            nombre_ca = dato_ca.datos_originales.get('nombre')
                            if nombre_ca:
                                curso_academico = CursoAcademico.objects.filter(nombre=nombre_ca).first()
                    
                    if not teacher or not curso_academico:
                        logger.warning(f"Faltan datos para curso {name}")
                        continue
                    
                    # Buscar o crear curso
                    curso, created = Curso.objects.get_or_create(
                        name=name,
                        curso_academico=curso_academico,
                        defaults={
                            'teacher': teacher,
                            'description': datos.get('description', ''),
                            'area': datos.get('area', 'idiomas'),
                            'tipo': datos.get('tipo', 'curso'),
                            'class_quantity': datos.get('class_quantity', 0),
                            'status': datos.get('status', 'F'),
                            'enrollment_deadline': datos.get('enrollment_deadline'),
                            'start_date': datos.get('start_date'),
                        }
                    )
                    
                    if not created:
                        # Actualizar campos si el curso ya existe
                        curso.teacher = teacher
                        curso.description = datos.get('description', curso.description)
                        curso.save()
                    
                    mapeo_cursos[dato.id_original] = curso
                    estadisticas['cursos_combinados'] += 1
                    logger.info(f"Curso {'creado' if created else 'actualizado'}: {name}")
                    
                except Exception as e:
                    error_msg = f"Error procesando curso {dato.id_original}: {str(e)}"
                    logger.error(error_msg)
                    errores.append(error_msg)
                    continue
            
            # 5. COMBINAR principal_matriculas
            logger.info("=== Iniciando combinación de principal_matriculas ===")
            datos_matriculas = DatoArchivadoDinamico.objects.filter(
                tabla_origen='principal_matriculas'
            )
            mapeo_matriculas = {}
            
            for dato in datos_matriculas:
                try:
                    datos = dato.datos_originales
                    course_id = datos.get('course_id')
                    student_id = datos.get('student_id')
                    
                    # Buscar curso y estudiante
                    curso = mapeo_cursos.get(course_id)
                    estudiante = mapeo_usuarios.get(student_id)
                    
                    if not curso or not estudiante:
                        logger.warning(f"Faltan datos para matrícula {dato.id_original}")
                        continue
                    
                    # Buscar o crear matrícula
                    matricula, created = Matriculas.objects.get_or_create(
                        course=curso,
                        student=estudiante,
                        defaults={
                            'activo': datos.get('activo', True),
                            'fecha_matricula': datos.get('fecha_matricula', date.today()),
                            'estado': datos.get('estado', 'P'),
                        }
                    )
                    
                    if not created:
                        # Actualizar si existe
                        matricula.activo = datos.get('activo', matricula.activo)
                        matricula.estado = datos.get('estado', matricula.estado)
                        matricula.save()
                    
                    mapeo_matriculas[dato.id_original] = matricula
                    estadisticas['matriculas_combinadas'] += 1
                    
                except Exception as e:
                    error_msg = f"Error procesando matrícula {dato.id_original}: {str(e)}"
                    logger.error(error_msg)
                    errores.append(error_msg)
                    continue
            
            # 6. COMBINAR principal_asistencia
            logger.info("=== Iniciando combinación de principal_asistencia ===")
            datos_asistencias = DatoArchivadoDinamico.objects.filter(
                tabla_origen='principal_asistencia'
            )
            
            for dato in datos_asistencias:
                try:
                    datos = dato.datos_originales
                    course_id = datos.get('course_id')
                    student_id = datos.get('student_id')
                    fecha = datos.get('date')
                    
                    curso = mapeo_cursos.get(course_id)
                    estudiante = mapeo_usuarios.get(student_id)
                    
                    if not curso or not estudiante or not fecha:
                        continue
                    
                    # Convertir fecha si es string
                    if isinstance(fecha, str):
                        try:
                            fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
                        except:
                            continue
                    
                    # Buscar o crear asistencia
                    asistencia, created = Asistencia.objects.get_or_create(
                        course=curso,
                        student=estudiante,
                        date=fecha,
                        defaults={
                            'presente': datos.get('presente', False)
                        }
                    )
                    
                    if not created:
                        asistencia.presente = datos.get('presente', asistencia.presente)
                        asistencia.save()
                    
                    estadisticas['asistencias_combinadas'] += 1
                    
                except Exception as e:
                    error_msg = f"Error procesando asistencia {dato.id_original}: {str(e)}"
                    logger.error(error_msg)
                    errores.append(error_msg)
                    continue
            
            # 7. COMBINAR principal_calificaciones
            logger.info("=== Iniciando combinación de principal_calificaciones ===")
            datos_calificaciones = DatoArchivadoDinamico.objects.filter(
                tabla_origen='principal_calificaciones'
            )
            mapeo_calificaciones = {}
            
            for dato in datos_calificaciones:
                try:
                    datos = dato.datos_originales
                    matricula_id = datos.get('matricula_id')
                    course_id = datos.get('course_id')
                    student_id = datos.get('student_id')
                    
                    matricula = mapeo_matriculas.get(matricula_id)
                    curso = mapeo_cursos.get(course_id)
                    estudiante = mapeo_usuarios.get(student_id)
                    
                    if not matricula or not curso or not estudiante:
                        continue
                    
                    # Buscar o crear calificación
                    calificacion, created = Calificaciones.objects.get_or_create(
                        matricula=matricula,
                        course=curso,
                        student=estudiante,
                        defaults={
                            'average': datos.get('average')
                        }
                    )
                    
                    if not created and datos.get('average'):
                        calificacion.average = datos.get('average')
                        calificacion.save()
                    
                    mapeo_calificaciones[dato.id_original] = calificacion
                    estadisticas['calificaciones_combinadas'] += 1
                    
                except Exception as e:
                    error_msg = f"Error procesando calificación {dato.id_original}: {str(e)}"
                    logger.error(error_msg)
                    errores.append(error_msg)
                    continue
            
            # 8. COMBINAR principal_notaindividual
            logger.info("=== Iniciando combinación de principal_notaindividual ===")
            datos_notas = DatoArchivadoDinamico.objects.filter(
                tabla_origen='principal_notaindividual'
            )
            
            for dato in datos_notas:
                try:
                    datos = dato.datos_originales
                    calificacion_id = datos.get('calificacion_id')
                    valor = datos.get('valor')
                    fecha_creacion = datos.get('fecha_creacion')
                    
                    calificacion = mapeo_calificaciones.get(calificacion_id)
                    
                    if not calificacion or not valor:
                        continue
                    
                    # Convertir fecha
                    if isinstance(fecha_creacion, str):
                        try:
                            fecha_creacion = datetime.strptime(fecha_creacion, '%Y-%m-%d').date()
                        except:
                            fecha_creacion = date.today()
                    
                    # Crear nota (no verificar duplicados para permitir múltiples notas)
                    nota = NotaIndividual.objects.create(
                        calificacion=calificacion,
                        valor=valor,
                        fecha_creacion=fecha_creacion or date.today()
                    )
                    
                    estadisticas['notas_combinadas'] += 1
                    
                except Exception as e:
                    error_msg = f"Error procesando nota {dato.id_original}: {str(e)}"
                    logger.error(error_msg)
                    errores.append(error_msg)
                    continue
            
            # 9. PROCESAR OTRAS TABLAS
            logger.info("=== Procesando otras tablas ===")
            tablas_procesadas = [
                'auth_user', 'Docencia_studentpersonalinformation', 'accounts_registro',
                'principal_cursoacademico', 'principal_curso', 'principal_matriculas',
                'principal_asistencia', 'principal_calificaciones', 'principal_notaindividual'
            ]
            
            otras_tablas = DatoArchivadoDinamico.objects.exclude(
                tabla_origen__in=tablas_procesadas
            ).values('tabla_origen').distinct()
            
            for tabla_info in otras_tablas:
                tabla_nombre = tabla_info['tabla_origen']
                count = DatoArchivadoDinamico.objects.filter(tabla_origen=tabla_nombre).count()
                logger.info(f"Tabla no procesada: {tabla_nombre} ({count} registros)")
                estadisticas['otras_tablas'] += count
        
        # Preparar respuesta
        resultado = {
            'success': True,
            **estadisticas,
            'campos_agregados': len(campos_agregados),
            'campos_nuevos_detectados': campos_agregados[:10],
            'errores_count': len(errores),
            'mensaje': f'Combinación completada exitosamente'
        }
        
        if errores:
            resultado['errores_muestra'] = errores[:5]
        
        logger.info(f"=== Combinación completada ===")
        logger.info(f"Estadísticas: {estadisticas}")
        return JsonResponse(resultado)
        
    except Exception as e:
        logger.error(f"Error general en combinación de datos: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        
        return JsonResponse({
            'success': False,
            'error': f'Error al combinar datos: {str(e)}',
            **{k: 0 for k in estadisticas.keys()}
        })
